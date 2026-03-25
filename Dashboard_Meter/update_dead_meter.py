#!/usr/bin/env python3
"""
อัปเดตข้อมูลมาตรตายใน index.html อัตโนมัติ
==================================================
เงื่อนไขการนับมาตรตาย:
  1. สภาพมาตร (col 12) = "มาตรไม่เดิน" เท่านั้น
  2. การเปลี่ยนมาตร (col 16) ≠ "เปลี่ยนแล้ว"
  3. เลขที่ผู้ใช้น้ำ (col 2) ไม่ซ้ำกัน ถ้าซ้ำนับ 1

แหล่งข้อมูล:
  - มาตรตาย: ข้อมูลดิบ/มาตรวัดน้ำผิดปกติ/METER_xxxx.xlsx
  - มาตรทั้งหมด (TOTAL_METERS): ผู้ใช้น้ำต้นงวด จาก OIS

วิธีใช้:
  python update_dead_meter.py
  python update_dead_meter.py --month 8   (ระบุคอลัมน์เดือนใน OIS เอง)
"""

import os
import re
import sys
import argparse
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERROR: ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("ERROR: ต้องติดตั้ง xlrd ก่อน: pip install xlrd")
    sys.exit(1)

# === CONFIG ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INDEX_HTML = os.path.join(SCRIPT_DIR, "index.html")
METER_DIR = os.path.join(SCRIPT_DIR, "ข้อมูลดิบ", "มาตรวัดน้ำผิดปกติ")
OIS_DIR = os.path.join(SCRIPT_DIR, "..", "Dashboard_Leak", "ข้อมูลดิบ", "OIS")

METER_SIZES = ["1/2", "3/4", "1", "1 1/2", "2", "2 1/2", "3", "4", "6", "8"]

BRANCH_MAP = {
    "1102": "ชลบุรี(พ)", "1103": "พัทยา(พ)", "1104": "บ้านบึง", "1105": "พนัสนิคม",
    "1106": "ศรีราชา", "1107": "แหลมฉบัง", "1108": "ฉะเชิงเทรา", "1109": "บางปะกง",
    "1110": "บางคล้า", "1111": "พนมสารคาม", "1112": "ระยอง", "1113": "บ้านฉาง",
    "1114": "ปากน้ำประแสร์", "1115": "จันทบุรี", "1116": "ขลุง", "1117": "ตราด",
    "1118": "คลองใหญ่", "1119": "สระแก้ว", "1120": "วัฒนานคร", "1121": "อรัญประเทศ",
    "1122": "ปราจีนบุรี", "1123": "กบินทร์บุรี"
}

BRANCH_ORDER = [
    "ชลบุรี(พ)", "พัทยา(พ)", "บ้านบึง", "พนัสนิคม", "ศรีราชา", "แหลมฉบัง",
    "ฉะเชิงเทรา", "บางปะกง", "บางคล้า", "พนมสารคาม",
    "ระยอง", "บ้านฉาง", "ปากน้ำประแสร์",
    "จันทบุรี", "ขลุง", "ตราด", "คลองใหญ่",
    "สระแก้ว", "วัฒนานคร", "อรัญประเทศ",
    "ปราจีนบุรี", "กบินทร์บุรี"
]

# OIS sheet name -> METER_BRANCHES name
OIS_SHEET_MAP = {
    'ป.ชลบุรี น.3': 'ชลบุรี(พ)', 'ป.บ้านบึง น.4': 'บ้านบึง',
    'ป.พนัสนิคม น.5': 'พนัสนิคม', 'ป.ศรีราชา น.6': 'ศรีราชา',
    'ป.แหลมฉบัง น.7': 'แหลมฉบัง', 'ป.พัทยา น.8': 'พัทยา(พ)',
    'ป.ฉะเชิงเทรา น.9': 'ฉะเชิงเทรา', 'ป.บางปะกง น.10': 'บางปะกง',
    'ป.บางคล้า น.11': 'บางคล้า', 'ป.พนมสารคาม น.12': 'พนมสารคาม',
    'ป.ระยอง น.13': 'ระยอง', 'ป.บ้านฉาง น.14': 'บ้านฉาง',
    'ป.ปากน้ำประแสร์ น.15': 'ปากน้ำประแสร์', 'ป.จันทบุรี น.16': 'จันทบุรี',
    'ป.ขลุง น.17': 'ขลุง', 'ป.ตราด น.18': 'ตราด',
    'ป.คลองใหญ่ น.19': 'คลองใหญ่', 'ป.สระแก้ว น.20': 'สระแก้ว',
    'ป.วัฒนา น.21': 'วัฒนานคร', 'ป.อรัญประเทศ น.22': 'อรัญประเทศ',
    'ป.ปราจีน น.23': 'ปราจีนบุรี', 'ป.กบินทร์ น.24': 'กบินทร์บุรี',
}

# OIS month column mapping (row 3 headers):
# col 5 = ต.ค., col 6 = พ.ย., col 7 = ธ.ค., col 8 = ม.ค., col 9 = ก.พ.,
# col 10 = มี.ค., col 11 = เม.ย., col 12 = พ.ค., col 13 = มิ.ย.,
# col 14 = ก.ค., col 15 = ส.ค., col 16 = ก.ย.


def normalize_size(s):
    """แปลงขนาดมาตรให้ตรงกับ METER_SIZES"""
    s = str(s).strip()
    if s in METER_SIZES:
        return s
    if 'ตั้งแต่ 8' in s or '8 นิ้ว' in s:
        return '8'
    clean = s.replace(' ', '')
    for ms in METER_SIZES:
        if clean == ms.replace(' ', ''):
            return ms
    return None


def extract_dead_meters():
    """ดึงข้อมูลมาตรตายจาก Excel ตามเงื่อนไข"""
    results = {}
    for fname in sorted(os.listdir(METER_DIR)):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        code = fname.replace('METER_', '').replace('.xlsx', '')
        branch = BRANCH_MAP.get(code)
        if not branch:
            continue

        try:
            wb = openpyxl.load_workbook(os.path.join(METER_DIR, fname), data_only=True)
        except Exception as e:
            print(f"  WARNING: ข้าม {fname}: {e}")
            results[branch] = {"total": 0, "sizes": {s: 0 for s in METER_SIZES}}
            continue

        ws = wb.active
        seen = set()
        sizes = Counter()
        total = 0

        for r in range(2, ws.max_row + 1):
            cid = ws.cell(row=r, column=2).value
            if cid is None:
                continue
            cid = str(cid).strip()
            if cid in seen:
                continue

            # เงื่อนไข 1: สภาพมาตร = "มาตรไม่เดิน"
            condition = ws.cell(row=r, column=12).value
            if condition is None or str(condition).strip() != "มาตรไม่เดิน":
                continue

            # เงื่อนไข 2: การเปลี่ยนมาตร ≠ "เปลี่ยนแล้ว"
            change = ws.cell(row=r, column=16).value
            if change is not None and str(change).strip() == "เปลี่ยนแล้ว":
                continue

            seen.add(cid)
            total += 1
            sv = ws.cell(row=r, column=9).value
            if sv is not None:
                ns = normalize_size(sv)
                if ns:
                    sizes[ns] += 1

        results[branch] = {"total": total, "sizes": {s: sizes.get(s, 0) for s in METER_SIZES}}
        wb.close()

    return results


def find_latest_ois():
    """หาไฟล์ OIS ปีล่าสุด"""
    ois_files = [f for f in os.listdir(OIS_DIR) if f.startswith('OIS_') and f.endswith('.xls')]
    if not ois_files:
        return None
    latest = sorted(ois_files)[-1]
    return os.path.join(OIS_DIR, latest)


def detect_month_col(ws):
    """หาคอลัมน์เดือนล่าสุดที่มีข้อมูลใน row 5 (ผู้ใช้น้ำต้นงวด)"""
    last_col = 5  # default ต.ค.
    for c in range(5, 17):  # col 5 (ต.ค.) to col 16 (ก.ย.)
        val = ws.cell_value(5, c)
        if val != '' and val != 0:
            last_col = c
    return last_col


def extract_total_meters(month_col=None):
    """ดึงจำนวนผู้ใช้น้ำต้นงวดจาก OIS"""
    ois_file = find_latest_ois()
    if not ois_file:
        print("ERROR: ไม่พบไฟล์ OIS ใน", OIS_DIR)
        return None, None

    print(f"  OIS file: {os.path.basename(ois_file)}")
    wb = xlrd.open_workbook(ois_file)

    # Auto-detect month if not specified
    if month_col is None:
        # Use first branch sheet to detect
        for sheet_name in OIS_SHEET_MAP:
            try:
                ws = wb.sheet_by_name(sheet_name)
                month_col = detect_month_col(ws)
                break
            except xlrd.biffh.XLRDError:
                continue

    # Get month name from header
    ws0 = wb.sheet_by_name(list(OIS_SHEET_MAP.keys())[0])
    month_name = ws0.cell_value(3, month_col) if month_col < ws0.ncols else f"col {month_col}"
    print(f"  เดือนที่ใช้: {month_name} (column {month_col})")

    results = {}
    for sheet_name, branch_name in OIS_SHEET_MAP.items():
        try:
            ws = wb.sheet_by_name(sheet_name)
        except xlrd.biffh.XLRDError:
            print(f"  WARNING: ไม่พบ sheet '{sheet_name}'")
            results[branch_name] = 0
            continue

        # Row 5 = ผู้ใช้น้ำต้นงวด
        val = ws.cell_value(5, month_col)
        results[branch_name] = int(val) if val != '' else 0

    return results, month_name


def generate_js_dead_meter(dead_data):
    """สร้าง JS code สำหรับ DEAD_METER"""
    lines = ["var DEAD_METER={"]
    for i, b in enumerate(BRANCH_ORDER):
        d = dead_data.get(b, {"total": 0, "sizes": {s: 0 for s in METER_SIZES}})
        sz = ",".join([f'"{s}":{d["sizes"].get(s, 0)}' for s in METER_SIZES])
        comma = "," if i < len(BRANCH_ORDER) - 1 else ""
        lines.append(f'"{b}":{{"total":{d["total"]},"sizes":{{{sz}}}}}{comma}')
    lines.append("};")
    return "\n".join(lines)


def generate_js_total_meters(total_data):
    """สร้าง JS code สำหรับ TOTAL_METERS"""
    lines = ["var TOTAL_METERS={"]
    groups = [BRANCH_ORDER[0:6], BRANCH_ORDER[6:10], BRANCH_ORDER[10:13],
              BRANCH_ORDER[13:17], BRANCH_ORDER[17:20], BRANCH_ORDER[20:22]]
    for gi, group in enumerate(groups):
        parts = []
        for b in group:
            parts.append(f'"{b}":{total_data.get(b, 0)}')
        comma = "," if gi < len(groups) - 1 else ""
        lines.append(",".join(parts) + comma)
    lines.append("};")
    return "\n".join(lines)


def update_index_html(dead_js, total_js):
    """อัปเดต DEAD_METER และ TOTAL_METERS ใน index.html"""
    with open(INDEX_HTML, 'r', encoding='utf-8') as f:
        content = f.read()

    # Replace DEAD_METER
    content = re.sub(
        r'var DEAD_METER=\{[\s\S]*?\};',
        dead_js,
        content,
        count=1
    )

    # Replace TOTAL_METERS
    content = re.sub(
        r'var TOTAL_METERS=\{[\s\S]*?\};',
        total_js,
        content,
        count=1
    )

    with open(INDEX_HTML, 'w', encoding='utf-8') as f:
        f.write(content)


def main():
    parser = argparse.ArgumentParser(description='อัปเดตข้อมูลมาตรตายใน index.html')
    parser.add_argument('--month', type=int, default=None,
                        help='คอลัมน์เดือนใน OIS (5=ต.ค., 6=พ.ย., 7=ธ.ค., 8=ม.ค., ...16=ก.ย.) ถ้าไม่ระบุจะหาอัตโนมัติ')
    args = parser.parse_args()

    print("=" * 60)
    print("อัปเดตข้อมูลมาตรตาย - Dashboard Meter")
    print("=" * 60)

    # 1. Extract dead meters
    print("\n[1/3] ดึงข้อมูลมาตรตายจาก Excel...")
    dead_data = extract_dead_meters()
    grand_dead = sum(d["total"] for d in dead_data.values())
    print(f"  รวมมาตรตาย: {grand_dead:,} ราย")

    # 2. Extract total meters from OIS
    print("\n[2/3] ดึงผู้ใช้น้ำต้นงวดจาก OIS...")
    total_data, month_name = extract_total_meters(args.month)
    if total_data is None:
        sys.exit(1)
    grand_total = sum(total_data.values())
    print(f"  รวมผู้ใช้น้ำต้นงวด: {grand_total:,} ราย")

    # 3. Update index.html
    print("\n[3/3] อัปเดต index.html...")
    dead_js = generate_js_dead_meter(dead_data)
    total_js = generate_js_total_meters(total_data)
    update_index_html(dead_js, total_js)
    print("  อัปเดตสำเร็จ!")

    # Summary
    print("\n" + "=" * 60)
    print(f"{'สาขา':<18} {'ผู้ใช้น้ำต้นงวด':>14} {'มาตรตาย':>10} {'%':>8}")
    print("-" * 54)
    for b in BRANCH_ORDER:
        d = dead_data.get(b, {"total": 0})["total"]
        t = total_data.get(b, 0)
        pct = (d / t * 100) if t > 0 else 0
        print(f"{b:<18} {t:>14,} {d:>10,} {pct:>7.2f}%")
    print("-" * 54)
    pct_all = (grand_dead / grand_total * 100) if grand_total > 0 else 0
    print(f"{'รวม กปภ.ข.1':<18} {grand_total:>14,} {grand_dead:>10,} {pct_all:>7.2f}%")
    print("=" * 60)


if __name__ == "__main__":
    main()
