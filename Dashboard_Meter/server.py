# -*- coding: utf-8 -*-
"""
Dashboard Meter — Local Development Server
============================================
Flask server สำหรับ Dashboard มาตรวัดน้ำ
รับ upload ไฟล์ Excel → parse ข้อมูลมาตรตายทันที → บันทึก data.json
index.html ดึงข้อมูลจาก /api/data

Usage:
    python server.py
    หรือ ดับเบิลคลิก start_server.bat

API Endpoints:
    GET  /                     → serve index.html หรือ manage.html
    GET  /api/ping             → health check
    GET  /api/data             → ข้อมูลมาตรตายทั้งหมด (JSON) + inventory ไฟล์
    POST /api/upload/<category> → upload ไฟล์ → parse → บันทึก data.json (ต้องส่ง data_date ด้วย)
    DELETE /api/data/<category>/<snapshot_date>/<filename> → ลบไฟล์ + อัปเดต data.json
    POST /api/open-folder      → เปิดโฟลเดอร์ ข้อมูลดิบ/ ใน File Explorer
    POST /api/open-main        → เปิด ../index.html (หน้า Landing)
"""

from flask import Flask, request, jsonify, send_from_directory
import os, sys, json, shutil, traceback, subprocess, re
from datetime import datetime
from collections import Counter
import platform

try:
    import openpyxl
except ImportError:
    print("WARNING: openpyxl ไม่ได้ติดตั้ง — pip install openpyxl")
    openpyxl = None

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DATA_DIR = os.path.join(BASE_DIR, 'ข้อมูลดิบ')
DATA_FILE = os.path.join(BASE_DIR, 'data.json')
PORT = 5003

METER_SIZES = ["1/2", "3/4", "1", "1 1/2", "2", "2 1/2", "3", "4", "6", "8"]

BRANCH_CODE_MAP = {
    "1102": "ชลบุรี(พ)", "1103": "พัทยา(พ)", "1104": "บ้านบึง", "1105": "พนัสนิคม",
    "1106": "ศรีราชา", "1107": "แหลมฉบัง", "1108": "ฉะเชิงเทรา", "1109": "บางปะกง",
    "1110": "บางคล้า", "1111": "พนมสารคาม", "1112": "ระยอง", "1113": "บ้านฉาง",
    "1114": "ปากน้ำประแสร์", "1115": "จันทบุรี", "1116": "ขลุง", "1117": "ตราด",
    "1118": "คลองใหญ่", "1119": "สระแก้ว", "1120": "วัฒนานคร", "1121": "อรัญประเทศ",
    "1122": "ปราจีนบุรี", "1123": "กบินทร์บุรี"
}

BRANCH_ORDER = [
    "ชลบุรี(พ)", "พัทยา(พ)", "บ้านบึง", "พนัสนิคม", "ศรีราชา", "แหลมฉบัง",
    "ฉะเชิงเทรา", "บางปะกง", "บางคล้า", "พนมสารคาม",
    "ระยอง", "บ้านฉาง", "ปากน้ำประแสร์",
    "จันทบุรี", "ขลุง", "ตราด", "คลองใหญ่",
    "สระแก้ว", "วัฒนานคร", "อรัญประเทศ",
    "ปราจีนบุรี", "กบินทร์บุรี"
]

# Thai month names for date formatting
TH_MONTHS = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
             'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']

# Category mapping: URL slug → Thai folder name
CATEGORY_MAP = {
    'abnormal': 'มาตรวัดน้ำผิดปกติ',
}

# ─── Data Persistence ─────────────────────────────────────────────────────────

def load_data():
    """โหลดข้อมูลจาก data.json"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Migrate old format to snapshots format
                dm = data.get('dead_meter', {})
                if 'snapshots' not in dm:
                    data['dead_meter'] = {"snapshots": {}, "latest": ""}
                return data
        except:
            pass
    return {"dead_meter": {"snapshots": {}, "latest": ""}}


def save_data(data):
    """บันทึกข้อมูลลง data.json"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def parse_date_key(date_str):
    """
    Parse date string in various formats to a date key (YYYY-MM-DD in Buddhist era)
    Accepts: "2569-01-16", "16/01/2569", "2569-1-16", etc.
    Returns: ("2569-01-16", "ณ วันที่ 16 มกราคม 2569") or (None, None) if invalid
    """
    date_str = date_str.strip()

    # Try YYYY-MM-DD format (Buddhist era)
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            date_key = f"{year:04d}-{month:02d}-{day:02d}"
            date_label = f"ณ วันที่ {day} {TH_MONTHS[month]} {year}"
            return date_key, date_label

    # Try DD/MM/YYYY format (Buddhist era)
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            date_key = f"{year:04d}-{month:02d}-{day:02d}"
            date_label = f"ณ วันที่ {day} {TH_MONTHS[month]} {year}"
            return date_key, date_label

    return None, None


# ─── Excel Parser ─────────────────────────────────────────────────────────────

def normalize_size(s):
    """แปลงขนาดมาตรให้ตรงกับ METER_SIZES"""
    s = str(s).strip()
    if s in METER_SIZES:
        return s
    clean = s.replace(' ', '')
    for ms in METER_SIZES:
        if clean == ms.replace(' ', ''):
            return ms
    if '8' in s and ('ตั้งแต่' in s or 'นิ้ว' in s):
        return '8'
    return None


def parse_dead_meter_file(file_path):
    """
    อ่านไฟล์ Excel มาตรวัดน้ำผิดปกติ แล้วนับมาตรตายตามเงื่อนไข:
      1. สภาพมาตร (col 12) = "มาตรไม่เดิน"
      2. การเปลี่ยนมาตร (col 16) ≠ "เปลี่ยนแล้ว"
      3. เลขที่ผู้ใช้น้ำ (col 2) ไม่ซ้ำกัน
    """
    if not openpyxl:
        raise Exception("openpyxl ไม่ได้ติดตั้ง")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    seen = set()
    sizes = Counter()
    total = 0

    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=2).value
        if cid is None:
            continue
        cid = str(cid).strip()
        if cid in seen:
            continue

        # เงื่อนไข 1: สภาพมาตร = "มาตรไม่เดิน"
        condition = ws.cell(row=r, column=12).value
        if condition is None or str(condition).strip() != "มาตรไม่เดิน":
            continue

        # เงื่อนไข 2: การเปลี่ยนมาตร ≠ "เปลี่ยนแล้ว"
        change = ws.cell(row=r, column=16).value
        if change is not None and str(change).strip() == "เปลี่ยนแล้ว":
            continue

        seen.add(cid)
        total += 1

        sv = ws.cell(row=r, column=9).value
        if sv is not None:
            ns = normalize_size(str(sv))
            if ns:
                sizes[ns] += 1

    # ดึงเดือนปีตั้งหนี้จาก col 1 (เช่น "256812" → "2568-12")
    billing_month = None
    for r in range(2, min(ws.max_row + 1, 20)):
        v = ws.cell(row=r, column=1).value
        if v:
            vs = str(v).strip()
            if len(vs) == 6 and vs.isdigit():
                billing_month = vs[:4] + "-" + vs[4:]
                break

    wb.close()
    return {
        "total": total,
        "sizes": {s: sizes.get(s, 0) for s in METER_SIZES},
        "billing_month": billing_month
    }


# Create directories
os.makedirs(RAW_DATA_DIR, exist_ok=True)
for folder in CATEGORY_MAP.values():
    os.makedirs(os.path.join(RAW_DATA_DIR, folder), exist_ok=True)

# ─── Flask App ────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder=BASE_DIR)

# ─── CORS Middleware ──────────────────────────────────────────────────────────

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

# ─── Serve Static Files ───────────────────────────────────────────────────────

@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/manage.html')
def serve_manage():
    return send_from_directory(BASE_DIR, 'manage.html')

# ─── API Endpoints ────────────────────────────────────────────────────────────

@app.route('/api/ping')
def api_ping():
    """ตรวจสอบว่า server ทำงานอยู่"""
    return jsonify({'ok': True, 'version': '2.0', 'timestamp': datetime.now().isoformat()})

@app.route('/api/data')
def api_get_data():
    """คืนข้อมูลมาตรตายทั้งหมด (snapshots) + inventory ไฟล์"""
    data = load_data()

    # สร้าง inventory ไฟล์
    inventory = {}
    for slug, thai_name in CATEGORY_MAP.items():
        folder_path = os.path.join(RAW_DATA_DIR, thai_name)
        files = []
        last_modified = None

        if os.path.exists(folder_path):
            try:
                for f in os.listdir(folder_path):
                    fp = os.path.join(folder_path, f)
                    if os.path.isfile(fp) and not f.startswith('.'):
                        files.append(f)
                        mtime = os.path.getmtime(fp)
                        if last_modified is None or mtime > last_modified:
                            last_modified = mtime
                files.sort()
            except:
                pass

        inventory[slug] = {
            'thai_name': thai_name,
            'files': files,
            'count': len(files),
            'last_modified': datetime.fromtimestamp(last_modified).strftime('%d/%m/%Y %H:%M') if last_modified else None
        }

    # Load notes
    notes_file = os.path.join(RAW_DATA_DIR, 'notes.json')
    notes = {}
    if os.path.exists(notes_file):
        try:
            with open(notes_file, 'r', encoding='utf-8') as nf:
                notes = json.load(nf)
        except (json.JSONDecodeError, IOError):
            pass

    return jsonify({
        'ok': True,
        'inventory': inventory,
        'dead_meter': data.get('dead_meter', {}),
        'notes': notes
    })

@app.route('/api/upload/<category>', methods=['POST', 'OPTIONS'])
def api_upload(category):
    """
    รับ upload ไฟล์ → บันทึก → parse ข้อมูลมาตรตายทันที → บันทึก data.json
    ต้องส่ง data_date (วันที่ดึงข้อมูลจาก CIS Support) มาด้วย เช่น "2569-01-16"
    """
    if request.method == 'OPTIONS':
        return '', 204

    if category not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': f'ไม่รู้จัก category: {category}'}), 400

    # รับวันที่จาก form data
    data_date = request.form.get('data_date', '').strip()
    if not data_date:
        return jsonify({'ok': False, 'error': 'กรุณาระบุวันที่ดึงข้อมูล (data_date)'}), 400

    date_key, date_label = parse_date_key(data_date)
    if not date_key:
        return jsonify({'ok': False, 'error': f'รูปแบบวันที่ไม่ถูกต้อง: {data_date} (ใช้ YYYY-MM-DD เช่น 2569-01-16)'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': 'ไม่ได้เลือกไฟล์'}), 400

    folder_path = os.path.join(RAW_DATA_DIR, CATEGORY_MAP[category])
    os.makedirs(folder_path, exist_ok=True)

    PREFIX_MAP = { 'abnormal': 'METER' }
    data = load_data()

    # Ensure snapshot exists for this date
    snapshots = data['dead_meter']['snapshots']
    if date_key not in snapshots:
        snapshots[date_key] = {
            "date_label": date_label,
            "data": {},
            "total_meters": {},
            "files": {}
        }
    snapshot = snapshots[date_key]

    results = []
    errors = []

    for f in files:
        filename = f.filename.strip()
        if not filename:
            continue

        try:
            prefix = PREFIX_MAP.get(category, category.upper())
            name_only = os.path.splitext(filename)[0]
            ext = os.path.splitext(filename)[1] or '.xlsx'
            # สร้างส่วนวันที่สำหรับชื่อไฟล์ เช่น "25690317"
            date_suffix = date_key.replace('-', '')  # "2569-03-17" → "25690317"
            # Meter: extract branch code (4 digits like 1102)
            m = re.search(r'(\d{4})', name_only)
            if m:
                new_name = f"{prefix}_{m.group(1)}_{date_suffix}{ext}"
            else:
                clean = re.sub(r'[^\w\-.]', '_', name_only).strip('_')[:30]
                new_name = f"{prefix}_{clean}_{date_suffix}{ext}"

            dest_path = os.path.join(folder_path, new_name)

            overwrite = os.path.exists(dest_path)
            f.save(dest_path)

            # Parse ข้อมูลมาตรตายทันที
            branch = None
            parsed = None
            if category == 'abnormal' and m:
                branch = BRANCH_CODE_MAP.get(m.group(1))
                if branch and openpyxl:
                    try:
                        parsed = parse_dead_meter_file(dest_path)
                        snapshot['data'][branch] = parsed
                        snapshot['files'][branch] = new_name
                    except Exception as pe:
                        errors.append({'filename': new_name, 'error': f'parse ล้มเหลว: {pe}'})

            results.append({
                'filename': new_name, 'original': filename,
                'status': 'overwrite' if overwrite else 'success',
                'message': f'{filename} → {new_name}',
                'branch': branch,
                'dead_count': parsed['total'] if parsed else None
            })
        except Exception as e:
            errors.append({
                'filename': filename,
                'error': str(e)
            })

    # อัปเดต latest snapshot
    data['dead_meter']['latest'] = date_key
    save_data(data)

    return jsonify({
        'ok': True,
        'category': category,
        'thai_name': CATEGORY_MAP[category],
        'date_key': date_key,
        'date_label': date_label,
        'results': results,
        'errors': errors,
        'dead_meter': data.get('dead_meter', {})
    })

@app.route('/api/data/<category>/<snapshot_date>/<filename>', methods=['DELETE', 'OPTIONS'])
def api_delete_file(category, snapshot_date, filename):
    """ลบไฟล์ + อัปเดต data.json (snapshot)"""
    if request.method == 'OPTIONS':
        return '', 204

    if category not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': f'ไม่รู้จัก category: {category}'}), 400

    folder_path = os.path.join(RAW_DATA_DIR, CATEGORY_MAP[category])
    file_path = os.path.join(folder_path, filename)

    # Safety check: ensure file is within folder
    if not os.path.abspath(file_path).startswith(os.path.abspath(folder_path)):
        return jsonify({'ok': False, 'error': 'ไม่อนุญาต'}), 403

    try:
        if os.path.exists(file_path):
            os.remove(file_path)

            # ลบข้อมูลสาขาที่เกี่ยวข้องออกจาก snapshot
            if category == 'abnormal':
                code_match = re.search(r'(\d{4})', filename)
                if code_match:
                    branch = BRANCH_CODE_MAP.get(code_match.group(1))
                    if branch:
                        data = load_data()
                        snapshots = data['dead_meter']['snapshots']
                        if snapshot_date in snapshots:
                            snap = snapshots[snapshot_date]
                            snap['data'].pop(branch, None)
                            snap['files'].pop(branch, None)
                            snap['total_meters'].pop(branch, None)
                            # ถ้า snapshot ว่างเปล่า ลบทิ้ง
                            if not snap['data']:
                                del snapshots[snapshot_date]
                                # อัปเดต latest
                                if data['dead_meter']['latest'] == snapshot_date:
                                    data['dead_meter']['latest'] = max(snapshots.keys()) if snapshots else ""
                            save_data(data)

            return jsonify({'ok': True, 'filename': filename, 'deleted': True})
        else:
            return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# Keep old delete route for backward compatibility
@app.route('/api/data/<category>/<filename>', methods=['DELETE', 'OPTIONS'])
def api_delete_file_compat(category, filename):
    """ลบไฟล์ (backward compatible — ลบจาก latest snapshot)"""
    if request.method == 'OPTIONS':
        return '', 204

    if category not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': f'ไม่รู้จัก category: {category}'}), 400

    folder_path = os.path.join(RAW_DATA_DIR, CATEGORY_MAP[category])
    file_path = os.path.join(folder_path, filename)

    if not os.path.abspath(file_path).startswith(os.path.abspath(folder_path)):
        return jsonify({'ok': False, 'error': 'ไม่อนุญาต'}), 403

    try:
        if os.path.exists(file_path):
            os.remove(file_path)

            if category == 'abnormal':
                code_match = re.search(r'(\d{4})', filename)
                if code_match:
                    branch = BRANCH_CODE_MAP.get(code_match.group(1))
                    if branch:
                        data = load_data()
                        # ลบจากทุก snapshot ที่มีไฟล์นี้
                        for sk, snap in data['dead_meter']['snapshots'].items():
                            if snap.get('files', {}).get(branch) == filename:
                                snap['data'].pop(branch, None)
                                snap['files'].pop(branch, None)
                                snap['total_meters'].pop(branch, None)
                        # ลบ snapshot ที่ว่างเปล่า
                        empty_keys = [k for k, v in data['dead_meter']['snapshots'].items() if not v.get('data')]
                        for k in empty_keys:
                            del data['dead_meter']['snapshots'][k]
                        if data['dead_meter']['latest'] in empty_keys:
                            snaps = data['dead_meter']['snapshots']
                            data['dead_meter']['latest'] = max(snaps.keys()) if snaps else ""
                        save_data(data)

            return jsonify({'ok': True, 'filename': filename, 'deleted': True})
        else:
            return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/notes/<slug>', methods=['POST', 'OPTIONS'])
def api_save_note(slug):
    """บันทึกช่วยจำสำหรับแต่ละหมวด"""
    if request.method == 'OPTIONS':
        return '', 204

    if slug not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': 'invalid slug'}), 400

    body = request.get_json(force=True)
    text = body.get('text', '')

    notes_file = os.path.join(RAW_DATA_DIR, 'notes.json')
    notes = {}
    if os.path.exists(notes_file):
        try:
            with open(notes_file, 'r', encoding='utf-8') as f:
                notes = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    notes[slug] = text
    with open(notes_file, 'w', encoding='utf-8') as f:
        json.dump(notes, f, ensure_ascii=False, indent=2)

    return jsonify({'ok': True})

@app.route('/api/open-folder', methods=['POST', 'OPTIONS'])
def api_open_folder():
    """เปิดโฟลเดอร์ ข้อมูลดิบ/ ใน File Explorer"""
    if request.method == 'OPTIONS':
        return '', 204

    try:
        if platform.system() == 'Windows':
            os.startfile(RAW_DATA_DIR)
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', RAW_DATA_DIR])
        else:
            subprocess.Popen(['xdg-open', RAW_DATA_DIR])

        return jsonify({'ok': True, 'path': RAW_DATA_DIR})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/open-main', methods=['POST', 'OPTIONS'])
def api_open_main():
    """เปิดหน้าหลัก (Landing Page) ใน browser"""
    if request.method == 'OPTIONS':
        return '', 204

    try:
        parent_dir = os.path.dirname(BASE_DIR)
        main_file = os.path.join(parent_dir, 'index.html')

        if platform.system() == 'Windows':
            os.startfile(main_file)
        else:
            subprocess.Popen(['xdg-open', main_file])

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ─── Catch-all Static Files (must be AFTER all API routes) ────────────────────

@app.route('/<path:path>')
def serve_static(path):
    full_path = os.path.join(BASE_DIR, path)
    if os.path.isfile(full_path):
        return send_from_directory(BASE_DIR, path)
    return send_from_directory(BASE_DIR, 'index.html')

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("=" * 60)
    print("  Dashboard Meter — Development Server v2.0")
    print(f"  http://localhost:{PORT}")
    print("=" * 60)
    print(f"  Base directory : {BASE_DIR}")
    print(f"  Raw data dir   : {RAW_DATA_DIR}")
    print(f"  Data file      : {DATA_FILE}")
    print()

    app.run(host='0.0.0.0', port=PORT, debug=True)
