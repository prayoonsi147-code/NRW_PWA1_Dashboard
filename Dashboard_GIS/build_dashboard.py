#!/usr/bin/env python3
"""
build_dashboard.py - สร้าง Dashboard แผนที่แนวท่อ (GIS) กปภ.เขต 1
อ่านข้อมูลจาก Excel แล้ว embed ลงใน index.html

ข้อมูลที่ฝัง:
  TAB 1: const DATA = {...}         — KPI จุดซ่อมท่อ (จาก ลงข้อมูลซ่อมท่อ/)
  TAB 2: const PRESSURE_DATA = {...} — แรงดันน้ำ (จาก แรงดันน้ำ/)
         const PRESSURE_MONTHS = [...]
  TAB 3: var PD1_DATA_FALLBACK = {...}  — กราฟค้างซ่อมเปรียบเทียบ (จาก ซ่อมท่อค้างระบบ/)
         var PD1_MONTHS_FALLBACK = [...]
         var PD2_DATA_FALLBACK = {...}  — กราฟค้างซ่อมสะสม
         var PD2_MONTHS_FALLBACK = [...]
         var PD3_FALLBACK = {...}       — ตารางงานค้างซ่อม
         var PENDING_UPDATE_DATE = '...'
         var PENDING_FY_LIST = [...]
"""
import openpyxl
import json
import os
import re
import calendar
from datetime import datetime
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DATA_DIR = os.path.join(SCRIPT_DIR, "ข้อมูลดิบ")
REPAIR_DIR = os.path.join(RAW_DATA_DIR, "ลงข้อมูลซ่อมท่อ")
PRESSURE_DIR = os.path.join(RAW_DATA_DIR, "แรงดันน้ำ")
PENDING_DIR = os.path.join(RAW_DATA_DIR, "ซ่อมท่อค้างระบบ")
HTML_TEMPLATE = os.path.join(SCRIPT_DIR, "index.html")

MONTH_NAMES = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.',
               'ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']

BRANCH_LIST = [
    "ชลบุรี","พัทยา","บ้านบึง","พนัสนิคม","ศรีราชา","แหลมฉบัง",
    "ฉะเชิงเทรา","บางปะกง","บางคล้า","พนมสารคาม","ระยอง","บ้านฉาง",
    "ปากน้ำประแสร์","จันทบุรี","ขลุง","ตราด","คลองใหญ่",
    "สระแก้ว","วัฒนานคร","อรัญประเทศ","ปราจีนบุรี","กบินทร์บุรี"
]

def clean_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    s = str(val).replace(',','').replace('\xa0','').replace(' ','').strip()
    if s == '' or s == '-': return 0
    try: return float(s)
    except: return 0

def parse_thai_date(val):
    """Parse วันที่ทั้ง datetime object และ string พ.ศ. เช่น '01/10/2568'"""
    if isinstance(val, datetime):
        by = val.year + 543 if val.year < 2500 else val.year
        return val, by
    if isinstance(val, str) and '/' in val:
        try:
            parts = val.strip().split('/')
            dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
            by = yyyy if yyyy > 2500 else yyyy + 543
            ce_year = by - 543
            dt = datetime(ce_year, mm, dd)
            return dt, by
        except Exception:
            return None, None
    return None, None

# ============================================================
#  TAB 1: KPI จุดซ่อมท่อ
# ============================================================

def read_repair_data():
    """Read repair xlsx files, group by month, pick latest per month"""
    if not os.path.isdir(REPAIR_DIR):
        print("  [SKIP] ไม่พบโฟลเดอร์ ลงข้อมูลซ่อมท่อ/")
        return None

    files = sorted([f for f in os.listdir(REPAIR_DIR) if f.endswith('.xlsx') and not f.startswith('~$')])
    print(f"  พบไฟล์ข้อมูล: {len(files)} ไฟล์")

    file_info = []
    for fname in files:
        m = re.search(r'(\d{6})', fname)
        if not m: continue
        digits = m.group(1)
        yy, mm, dd = int(digits[:2]), int(digits[2:4]), int(digits[4:6])
        month_key = f"{yy:02d}-{mm:02d}"
        file_info.append({'fname': fname, 'yy': yy, 'mm': mm, 'dd': dd, 'month_key': month_key})

    month_files = {}
    for fi in file_info:
        mk = fi['month_key']
        if mk not in month_files or fi['dd'] > month_files[mk]['dd']:
            month_files[mk] = fi

    print(f"  เดือนที่มีข้อมูล: {len(month_files)}")
    for mk in sorted(month_files.keys()):
        print(f"    {mk} <- {month_files[mk]['fname']}")

    all_data = {}
    branches_order = []
    for mk in sorted(month_files.keys()):
        fi = month_files[mk]
        fpath = os.path.join(REPAIR_DIR, fi['fname'])
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"  [WARNING] ข้ามไฟล์เสีย: {fi['fname']} ({e})")
            continue
        ws = wb[wb.sheetnames[0]]
        month_data = {}
        for r in range(2, ws.max_row + 1):
            branch = ws.cell(row=r, column=1).value
            if not branch or not isinstance(branch, str): continue
            branch = branch.strip()
            if branch in ('', 'ชื่อสาขา'): continue
            month_data[branch] = {
                'closed': clean_num(ws.cell(row=r, column=2).value),
                'complete': clean_num(ws.cell(row=r, column=3).value),
                'score': clean_num(ws.cell(row=r, column=4).value)
            }
            if branch not in branches_order:
                branches_order.append(branch)
        all_data[mk] = month_data
        wb.close()

    months_sorted = sorted(all_data.keys())
    if months_sorted:
        print(f"  ช่วงเดือน: {months_sorted[0]} - {months_sorted[-1]}")
    print(f"  จำนวนสาขา: {len(branches_order)}")

    return {
        'months': months_sorted,
        'branches': branches_order,
        'data': all_data,
        'month_names': {f"{i:02d}": MONTH_NAMES[i] for i in range(1, 13)}
    }

# ============================================================
#  TAB 2: แรงดันน้ำ
# ============================================================

def read_pressure_data():
    """Read PRESSURE_*.xlsx files → compute average per branch per month"""
    if not os.path.isdir(PRESSURE_DIR):
        print("  [SKIP] ไม่พบโฟลเดอร์ แรงดันน้ำ/")
        return None, None

    files = [f for f in os.listdir(PRESSURE_DIR)
             if f.startswith('PRESSURE_') and f.endswith('.xlsx') and not f.startswith('~$')]
    if not files:
        print("  [SKIP] ไม่พบไฟล์ PRESSURE_*.xlsx")
        return None, None

    print(f"  พบไฟล์แรงดัน: {len(files)} ไฟล์")

    # Month header pattern: "ต.ค. 68", "พ.ย. 68", etc.
    thai_month_map = {
        'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4, 'พ.ค.': 5, 'มิ.ย.': 6,
        'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12
    }

    # pressure_data: {month_key: {branch: avg_value}}
    pressure_data = {}
    all_months = set()

    for fname in sorted(files):
        # Extract branch name from filename: PRESSURE_ชลบุรี_ปีงบ69.xlsx
        m = re.match(r'PRESSURE_(.+?)_ปีงบ\d+\.xlsx', fname)
        if not m:
            m = re.match(r'PRESSURE_(.+?)\.xlsx', fname)
        if not m:
            continue
        branch_name = m.group(1)

        fpath = os.path.join(PRESSURE_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            ws = wb.active

            # Row 5 has month headers starting from col 6
            # Find month columns
            month_cols = {}  # {col_index: month_key}
            for c in range(1, ws.max_column + 1):
                header = ws.cell(5, c).value
                if not header or not isinstance(header, str):
                    continue
                header = header.strip()
                for thai_m, m_num in thai_month_map.items():
                    if header.startswith(thai_m):
                        # Extract year: "ต.ค. 68" → yy=68
                        yy_match = re.search(r'(\d{2})', header.replace(thai_m, ''))
                        if yy_match:
                            yy = int(yy_match.group(1))
                            mk = f"{yy:02d}-{m_num:02d}"
                            month_cols[c] = mk
                        break

            # Read data rows (7 to max_row), compute average per month
            for mk in month_cols.values():
                all_months.add(mk)

            for col_idx, mk in month_cols.items():
                total = 0.0
                count = 0
                for r in range(7, ws.max_row + 1):
                    v = ws.cell(r, col_idx).value
                    if isinstance(v, (int, float)) and v > 0:
                        total += v
                        count += 1
                if count > 0:
                    avg = round(total / count, 2)
                    if mk not in pressure_data:
                        pressure_data[mk] = {}
                    pressure_data[mk][branch_name] = avg

            wb.close()
        except Exception as e:
            print(f"  [WARNING] ข้ามไฟล์: {fname} ({e})")

    # Filter months that actually have non-zero data
    months_sorted = sorted(mk for mk in all_months
                          if mk in pressure_data and pressure_data[mk]
                          and any(v > 0 for v in pressure_data[mk].values()))
    print(f"  เดือนแรงดัน: {months_sorted}")
    print(f"  สาขาที่มีข้อมูล: {len(files)}")

    return pressure_data, months_sorted

# ============================================================
#  TAB 3: งานค้างซ่อม
# ============================================================

def read_pending_data():
    """Read pending repair xlsx → compute pd1, pd2, pd3 data for ALL fiscal years"""
    if not os.path.isdir(PENDING_DIR):
        print("  [SKIP] ไม่พบโฟลเดอร์ ซ่อมท่อค้างระบบ/")
        return None

    # Scan files — extract fiscal year from filename
    fy_files = {}
    for fname in os.listdir(PENDING_DIR):
        if not fname.lower().endswith(('.xlsx', '.xls')) or fname.startswith('~$'):
            continue
        if fname.endswith('.cache.json'):
            continue
        fpath = os.path.join(PENDING_DIR, fname)
        m = re.search(r'(\d{1,2})-(\d{2})_to_(\d{1,2})-(\d{2})', fname)
        if m:
            start_mm, start_yy = int(m.group(1)), int(m.group(2))
            fy_be = 2500 + start_yy + 1 if start_mm >= 10 else 2500 + start_yy
            fy_files[fy_be] = fpath
            print(f"    ปีงบฯ {fy_be} <- {fname}")
        else:
            fy_files.setdefault(0, fpath)
            print(f"    (ไม่ระบุปี) <- {fname}")

    if not fy_files:
        print("  [SKIP] ไม่พบไฟล์ข้อมูลค้างซ่อม")
        return None

    fy_list = sorted([k for k in fy_files.keys() if k > 0])

    # Column indices (0-based)
    col_date = 3       # วันที่แจ้ง
    col_finish = 5     # วันเวลาเสร็จสิ้น
    col_branch = 19    # สาขา
    col_status = 26    # สถานะ
    data_start = 8     # row 9 = index 8

    all_fy_results = {}

    for fy in (fy_list if fy_list else [0]):
        fpath = fy_files.get(fy, list(fy_files.values())[0])
        if not os.path.isfile(fpath):
            continue

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            ws = wb.active
            all_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            print(f"  [WARNING] อ่านไฟล์ไม่ได้: {fpath} ({e})")
            continue

        fy_be = fy if fy > 0 else 2569
        fy_ce = fy_be - 543
        count_start = datetime(fy_ce, 1, 1)

        # --- Read all records ---
        records = []
        last_report_dt = None

        for row in all_rows[data_start:]:
            if len(row) <= col_status:
                continue
            date_val = row[col_date]
            if not date_val:
                continue
            dt, by = parse_thai_date(date_val)
            if dt is None:
                continue
            if last_report_dt is None or dt > last_report_dt:
                last_report_dt = dt

            finish_val = row[col_finish]
            finish_dt = None
            if finish_val:
                fdt, _ = parse_thai_date(finish_val)
                if fdt:
                    finish_dt = fdt
                elif isinstance(finish_val, str):
                    try:
                        finish_dt, _ = parse_thai_date(finish_val[:10])
                    except:
                        pass

            status = str(row[col_status] or '')
            branch = str(row[col_branch] or '').strip()
            if not branch:
                continue
            records.append({'dt': dt, 'by': by, 'finish_dt': finish_dt,
                           'status': status, 'branch': branch})

        # Build update_date
        if last_report_dt:
            by_lrd = last_report_dt.year + 543 if last_report_dt.year < 2500 else last_report_dt.year
            update_date = f"{last_report_dt.day:02d}-{last_report_dt.month:02d}-{by_lrd % 100}"
        else:
            update_date = ''

        # --- PD2: ค้างซ่อม ณ สิ้นเดือน (snapshot) ---
        month_set = set()
        for rec in records:
            if rec['dt'] >= count_start:
                month_set.add((rec['dt'].year, rec['dt'].month))

        sorted_months = sorted(month_set)
        pd2_months = []
        for y, m in sorted_months:
            yy = (y + 543) % 100 if y < 2500 else y % 100
            pd2_months.append(f"{yy:02d}-{m:02d}")

        pd2_data = {}
        for y, m in sorted_months:
            end_day = calendar.monthrange(y, m)[1]
            end_of_month = datetime(y, m, end_day, 23, 59, 59)
            yy = (y + 543) % 100 if y < 2500 else y % 100
            mk = f"{yy:02d}-{m:02d}"
            branch_counts = defaultdict(int)
            for rec in records:
                if rec['dt'] < count_start or rec['dt'] > end_of_month:
                    continue
                is_pending = False
                if rec['finish_dt'] and rec['finish_dt'] > end_of_month:
                    is_pending = True
                elif 'ซ่อมไม่เสร็จ' in rec['status']:
                    is_pending = True
                if is_pending:
                    branch_counts[rec['branch']] += 1
            pd2_data[mk] = dict(branch_counts)

        # --- PD1: เปรียบเทียบเดือน (derived from PD2) ---
        pd1_data = {}
        for i, mk in enumerate(pd2_months):
            prev_mk = pd2_months[i - 1] if i > 0 else None
            prev_snap = pd2_data.get(prev_mk, {}) if prev_mk else {}
            curr_snap = pd2_data.get(mk, {})
            branch_pairs = {}
            for b in BRANCH_LIST:
                pv = prev_snap.get(b, 0)
                cv = curr_snap.get(b, 0)
                branch_pairs[b] = [pv, cv]
            pd1_data[mk] = branch_pairs

        # --- PD3: ตารางงานซ่อมที่ยังไม่ปิดในระบบ (สถานะ = ซ่อมไม่เสร็จ) ---
        fy_yy_start = (fy_be - 2500 - 1) if fy_be > 0 else 68
        fy_yy_end = fy_yy_start + 1
        fy_months = []
        for mm in [10, 11, 12]:
            fy_months.append(f"{fy_yy_start:02d}-{mm:02d}")
        for mm in range(1, 10):
            fy_months.append(f"{fy_yy_end:02d}-{mm:02d}")

        pd3_data = defaultdict(lambda: defaultdict(int))
        for rec in records:
            if 'ซ่อมไม่เสร็จ' not in rec['status']:
                continue
            yy = rec['by'] % 100
            mk = f"{yy:02d}-{rec['dt'].month:02d}"
            if mk in fy_months:
                pd3_data[rec['branch']][mk] += 1

        # Convert defaultdict to regular dict
        pd3_data_clean = {}
        for branch, months in pd3_data.items():
            pd3_data_clean[branch] = dict(months)

        # Col totals + grand total
        col_totals = {mk: 0 for mk in fy_months}
        grand_total = 0
        for branch_data in pd3_data_clean.values():
            for mk, v in branch_data.items():
                col_totals[mk] = col_totals.get(mk, 0) + v
                grand_total += v

        print(f"  ปีงบฯ {fy_be}: records={len(records)}, pd2_months={pd2_months}, update={update_date}")

        all_fy_results[str(fy_be)] = {
            'update_date': update_date,
            'pd1_data': pd1_data,
            'pd1_months': pd2_months,
            'pd2_data': pd2_data,
            'pd2_months': pd2_months,
            'pd3': {
                'months': fy_months,
                'update_date': update_date,
                'data': pd3_data_clean,
                'col_totals': col_totals,
                'grand_total': grand_total
            }
        }

    return {
        'fy_list': fy_list,
        'results': all_fy_results
    }

# ============================================================
#  Embed into index.html
# ============================================================

def replace_js_var(html, pattern, new_value):
    """Replace a JS variable declaration in HTML using regex"""
    new_html, count = re.subn(pattern, new_value, html, count=1, flags=re.MULTILINE)
    return new_html if count > 0 else html

def embed_all(html, repair_data, pressure_data, pressure_months, pending_result):
    """Embed all data into HTML"""
    changes = []

    # --- TAB 1: KPI ---
    if repair_data:
        data_json = json.dumps(repair_data, ensure_ascii=False)
        if 'GIS_DATA_PLACEHOLDER' in html:
            html = html.replace('GIS_DATA_PLACEHOLDER', data_json)
            changes.append("TAB 1 KPI (placeholder)")
        else:
            pat = r'^const DATA = \{.*\};$'
            new_val = 'const DATA = ' + data_json + ';'
            html = replace_js_var(html, pat, new_val)
            changes.append("TAB 1 KPI (const DATA)")

    # --- TAB 2: Pressure ---
    if pressure_data and pressure_months:
        pdata_json = json.dumps(pressure_data, ensure_ascii=False)
        pmonths_json = json.dumps(pressure_months, ensure_ascii=False)
        html = replace_js_var(html,
            r'^const PRESSURE_DATA=\{.*\};$',
            'const PRESSURE_DATA=' + pdata_json + ';')
        html = replace_js_var(html,
            r'^const PRESSURE_MONTHS=\[.*\];$',
            'const PRESSURE_MONTHS=' + pmonths_json + ';')
        changes.append(f"TAB 2 Pressure ({len(pressure_months)} เดือน)")

    # --- TAB 3: Pending ---
    if pending_result:
        fy_list = pending_result['fy_list']
        results = pending_result['results']

        # Use latest FY
        latest_fy = str(fy_list[-1]) if fy_list else '2569'
        latest = results.get(latest_fy, {})

        if latest:
            update_date = latest['update_date']
            pd1_data = latest['pd1_data']
            pd1_months = latest['pd1_months']
            pd2_data = latest['pd2_data']
            pd2_months = latest['pd2_months']

            # PENDING_UPDATE_DATE (may be const or var)
            html = replace_js_var(html,
                r"^(?:const|var) PENDING_UPDATE_DATE='[^']*';$",
                f"const PENDING_UPDATE_DATE='{update_date}';")

            # PENDING_FY_LIST
            html = replace_js_var(html,
                r'^var PENDING_FY_LIST=\[.*\];.*$',
                'var PENDING_FY_LIST=' + json.dumps([int(x) for x in fy_list], ensure_ascii=False) + '; // fallback')

            # PD1
            html = replace_js_var(html,
                r'^var PD1_DATA_FALLBACK=\{.*\};$',
                'var PD1_DATA_FALLBACK=' + json.dumps(pd1_data, ensure_ascii=False) + ';')
            html = replace_js_var(html,
                r'^var PD1_MONTHS_FALLBACK=\[.*\];$',
                'var PD1_MONTHS_FALLBACK=' + json.dumps(pd1_months, ensure_ascii=False) + ';')

            # PD2
            html = replace_js_var(html,
                r'^var PD2_DATA_FALLBACK=\{.*\};$',
                'var PD2_DATA_FALLBACK=' + json.dumps(pd2_data, ensure_ascii=False) + ';')
            html = replace_js_var(html,
                r'^var PD2_MONTHS_FALLBACK=\[.*\];$',
                'var PD2_MONTHS_FALLBACK=' + json.dumps(pd2_months, ensure_ascii=False) + ';')

            # PD3_FALLBACK — embed all FYs
            pd3_fallback = {}
            for fy_key, fy_data in results.items():
                pd3_fallback[fy_key] = fy_data['pd3']
            # Replace the multi-line PD3_FALLBACK block
            pd3_json = json.dumps(pd3_fallback, ensure_ascii=False)
            # PD3_FALLBACK spans multiple lines — use DOTALL
            html_new = re.sub(
                r'var PD3_FALLBACK=\{.*?\n\};',
                'var PD3_FALLBACK=' + pd3_json + ';',
                html, count=1, flags=re.DOTALL)
            if html_new != html:
                html = html_new
            else:
                # Try single-line pattern
                html = replace_js_var(html,
                    r'^var PD3_FALLBACK=\{.*\};$',
                    'var PD3_FALLBACK=' + pd3_json + ';')

            changes.append(f"TAB 3 Pending (ปีงบฯ {','.join(str(x) for x in fy_list)}, update={update_date})")

    return html, changes

# ============================================================
#  Main Build
# ============================================================

def build():
    print("=" * 50)
    print("  Build Dashboard แผนที่แนวท่อ (GIS) กปภ.เขต 1")
    print("=" * 50)

    # --- 1) Read repair data (TAB 1) ---
    print("\n[1/4] อ่านข้อมูล KPI จุดซ่อมท่อ...")
    repair_data = read_repair_data()

    # --- 2) Read pressure data (TAB 2) ---
    print("\n[2/4] อ่านข้อมูลแรงดันน้ำ...")
    pressure_data, pressure_months = read_pressure_data()

    # --- 3) Read pending data (TAB 3) ---
    print("\n[3/4] อ่านข้อมูลงานค้างซ่อม...")
    pending_result = read_pending_data()

    # --- 4) Embed into index.html ---
    print("\n[4/4] Embed ข้อมูลลงใน index.html...")
    with open(HTML_TEMPLATE, 'r', encoding='utf-8') as f:
        html = f.read()

    html, changes = embed_all(html, repair_data, pressure_data, pressure_months, pending_result)

    for c in changes:
        print(f"  ✓ {c}")

    if not changes:
        print("  [WARNING] ไม่มีข้อมูลที่จะ embed!")

    output_path = os.path.join(SCRIPT_DIR, "index.html")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(output_path) / 1024
    print(f"\n  บันทึก: {output_path}")
    print(f"  ขนาดไฟล์: {size_kb:.1f} KB")
    print("\n  เสร็จสิ้น!")

if __name__ == '__main__':
    build()
