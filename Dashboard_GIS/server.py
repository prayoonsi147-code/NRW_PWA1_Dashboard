# -*- coding: utf-8 -*-
"""
Dashboard GIS — Local Development Server
=========================================
Flask server สำหรับ Dashboard แผนที่แนวท่อ
รับ upload ไฟล์ Excel → จัดเก็บลงโฟลเดอร์ ข้อมูลดิบ/ → ไม่ parse (ให้ build_dashboard.py จัดการ)

Usage:
    python server.py
    หรือ ดับเบิลคลิก start_server.bat

API Endpoints:
    GET  /                     → serve index.html หรือ manage.html
    GET  /api/ping             → health check
    GET  /api/data             → inventory ของไฟล์ข้อมูลดิบในแต่ละ category
    POST /api/upload/<category> → upload ไฟล์ไปยัง category folder
    DELETE /api/data/<category>/<filename> → ลบไฟล์เฉพาะ
    POST /api/open-folder      → เปิดโฟลเดอร์ ข้อมูลดิบ/ ใน File Explorer
    POST /api/open-main        → เปิด ../index.html (หน้า Landing)
    POST /api/rebuild          → รัน build_dashboard.py
"""

from flask import Flask, request, jsonify, send_from_directory
import os, sys, json, shutil, traceback, subprocess, threading, time
from datetime import datetime
import platform

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DATA_DIR = os.path.join(BASE_DIR, 'ข้อมูลดิบ')
PORT = 5002

# Category mapping: URL slug → Thai folder name
CATEGORY_MAP = {
    'repair': 'ลงข้อมูลซ่อมท่อ',
    'pressure': 'แรงดันน้ำ',
    'pending': 'ซ่อมท่อค้างระบบ',
}

# Create directories
os.makedirs(RAW_DATA_DIR, exist_ok=True)
for folder in CATEGORY_MAP.values():
    os.makedirs(os.path.join(RAW_DATA_DIR, folder), exist_ok=True)

# ─── Excel File Cache ─────────────────────────────────────────────────────────
# Cache openpyxl workbook data to avoid re-reading the same Excel file
# on every API call (pending-chart, pending-table, pending-detail all read the same file)

_excel_cache = {}  # { filepath: { 'rows': [...], 'mtime': float, 'time': float } }
_excel_lock = threading.Lock()
CACHE_TTL = 60  # seconds

def parse_thai_date(val):
    """Parse วันที่ทั้ง datetime object และ string พ.ศ. เช่น '01/10/2568'
    Returns: (datetime_obj, buddhist_year) or (None, None)"""
    if isinstance(val, datetime):
        by = val.year + 543 if val.year < 2500 else val.year
        return val, by
    if isinstance(val, str) and '/' in val:
        try:
            parts = val.strip().split('/')
            dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
            by = yyyy if yyyy > 2500 else yyyy + 543
            ce_year = by - 543
            dt = datetime(ce_year, mm, dd)
            return dt, by
        except Exception:
            return None, None
    return None, None

def _serialize_val(v):
    """แปลง datetime → dict สำหรับ JSON"""
    if isinstance(v, datetime):
        return {'__dt__': v.isoformat()}
    return v

def _deserialize_val(v):
    """แปลง dict → datetime กลับ"""
    if isinstance(v, dict) and '__dt__' in v:
        return datetime.fromisoformat(v['__dt__'])
    return v

def _build_json_cache(xlsx_path):
    """อ่าน Excel แล้วสร้าง .cache.json (เรียกตอน startup + upload)"""
    import openpyxl
    mtime = os.path.getmtime(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    json_path = xlsx_path + '.cache.json'
    json_rows = [[_serialize_val(c) for c in row] for row in rows]
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({'mtime': mtime, 'rows': json_rows}, f, ensure_ascii=False)

    with _excel_lock:
        _excel_cache[xlsx_path] = {'rows': rows, 'mtime': mtime, 'time': time.time()}
    return rows

def get_pending_rows(fpath):
    """อ่านข้อมูลค้างซ่อม — จาก memory → JSON cache → Excel (fallback)"""
    mtime = os.path.getmtime(fpath)

    # 1) Memory
    with _excel_lock:
        cached = _excel_cache.get(fpath)
        if cached and cached['mtime'] == mtime and (time.time() - cached['time']) < CACHE_TTL:
            return cached['rows']

    # 2) JSON disk cache
    json_path = fpath + '.cache.json'
    if os.path.exists(json_path):
        try:
            data = json.load(open(json_path, 'r', encoding='utf-8'))
            if data.get('mtime') == mtime:
                rows = [tuple(_deserialize_val(c) for c in r) for r in data['rows']]
                with _excel_lock:
                    _excel_cache[fpath] = {'rows': rows, 'mtime': mtime, 'time': time.time()}
                return rows
        except Exception:
            pass

    # 3) Fallback: อ่าน Excel แล้วสร้าง cache
    return _build_json_cache(fpath)

def preload_pending_cache():
    """เรียกตอน server เริ่ม — สร้าง JSON cache ล่วงหน้าสำหรับไฟล์ที่มีอยู่แล้ว"""
    import re as _re
    pending_dir = os.path.join(RAW_DATA_DIR, CATEGORY_MAP['pending'])
    if not os.path.isdir(pending_dir):
        return
    for fname in os.listdir(pending_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')) or fname.startswith('~$'):
            continue
        fpath = os.path.join(pending_dir, fname)
        json_path = fpath + '.cache.json'
        # สร้าง cache เฉพาะเมื่อยังไม่มี หรือ Excel เปลี่ยน
        needs_build = True
        if os.path.exists(json_path):
            try:
                data = json.load(open(json_path, 'r', encoding='utf-8'))
                if data.get('mtime') == os.path.getmtime(fpath):
                    needs_build = False
            except Exception:
                pass
        if needs_build:
            try:
                print(f'  Building cache for {fname}...')
                _build_json_cache(fpath)
                print(f'  ✓ Cache ready: {fname}')
            except Exception as e:
                print(f'  ✗ Cache failed: {fname} — {e}')

# ─── Flask App ────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder=BASE_DIR)

# ─── CORS Middleware ──────────────────────────────────────────────────────────

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

# ─── Serve Static Files ───────────────────────────────────────────────────────

@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/manage.html')
def serve_manage():
    return send_from_directory(BASE_DIR, 'manage.html')

# ─── API Endpoints ────────────────────────────────────────────────────────────

@app.route('/api/ping')
def api_ping():
    """ตรวจสอบว่า server ทำงานอยู่"""
    return jsonify({'ok': True, 'version': '1.0', 'timestamp': datetime.now().isoformat()})

@app.route('/api/data')
def api_get_data():
    """คืนรายชื่อไฟล์ในแต่ละ category folder"""
    inventory = {}

    for slug, thai_name in CATEGORY_MAP.items():
        folder_path = os.path.join(RAW_DATA_DIR, thai_name)
        files = []
        last_modified = None

        if os.path.exists(folder_path):
            try:
                for f in os.listdir(folder_path):
                    fp = os.path.join(folder_path, f)
                    if os.path.isfile(fp) and not f.startswith('.'):
                        files.append(f)
                        mtime = os.path.getmtime(fp)
                        if last_modified is None or mtime > last_modified:
                            last_modified = mtime
                files.sort()
            except Exception as e:
                pass

        inventory[slug] = {
            'thai_name': thai_name,
            'files': files,
            'count': len(files),
            'last_modified': datetime.fromtimestamp(last_modified).strftime('%d/%m/%Y %H:%M') if last_modified else None
        }

    # Load saved notes
    notes_file = os.path.join(RAW_DATA_DIR, 'notes.json')
    notes = {}
    if os.path.exists(notes_file):
        try:
            with open(notes_file, 'r', encoding='utf-8') as f:
                notes = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    return jsonify({'ok': True, 'inventory': inventory, 'notes': notes})

@app.route('/api/notes/<slug>', methods=['POST'])
def api_save_note(slug):
    """บันทึก note ของแต่ละ category"""
    if slug not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': 'invalid slug'}), 400
    body = request.get_json(force=True)
    text = body.get('text', '')
    notes_file = os.path.join(RAW_DATA_DIR, 'notes.json')
    notes = {}
    if os.path.exists(notes_file):
        try:
            with open(notes_file, 'r', encoding='utf-8') as f:
                notes = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    notes[slug] = text
    with open(notes_file, 'w', encoding='utf-8') as f:
        json.dump(notes, f, ensure_ascii=False, indent=2)
    return jsonify({'ok': True})

@app.route('/api/upload/<category>', methods=['POST', 'OPTIONS'])
def api_upload(category):
    """รับ upload ไฟล์ไปยัง category folder"""
    if request.method == 'OPTIONS':
        return '', 204

    if category not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': f'ไม่รู้จัก category: {category}'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': 'ไม่ได้เลือกไฟล์'}), 400

    folder_path = os.path.join(RAW_DATA_DIR, CATEGORY_MAP[category])
    os.makedirs(folder_path, exist_ok=True)

    PREFIX_MAP = { 'repair': 'GIS', 'pressure': 'PRESSURE', 'pending': 'PENDING' }
    import re

    results = []
    errors = []
    _pending_batch = []  # สำหรับ pending: เก็บไฟล์ทั้งหมดไว้รวมทีหลัง

    for f in files:
        filename = f.filename.strip()
        if not filename:
            continue

        try:
            prefix = PREFIX_MAP.get(category, category.upper())
            name_only = os.path.splitext(filename)[0]
            ext = os.path.splitext(filename)[1] or '.xlsx'

            # ── Category-specific auto-rename ──
            if category == 'repair':
                # repair: GIS_YYMMDD.xlsx — ดึงตัวเลข 6 หลัก (YYMMDD)
                m = re.search(r'(\d{6})', name_only)
                if m:
                    new_name = f"{prefix}_{m.group(1)}{ext}"
                else:
                    # fallback: ใช้วันที่ upload
                    today = datetime.now().strftime('%y%m%d')
                    new_name = f"{prefix}_{today}{ext}"

            elif category == 'pressure':
                # pressure: PRESSURE_สาขา_ปีงบYYYY.xlsx
                # ดึงชื่อสาขาจากชื่อไฟล์ + ปีงบประมาณจาก Row 3 ในไฟล์
                thai_parts = re.findall(r'[\u0e00-\u0e7f]+', name_only)
                branch_name = thai_parts[-1] if thai_parts else 'unknown'

                # อ่านปีงบประมาณจากเนื้อไฟล์ (Row 3: "ประจำปีงบประมาณ 25XX")
                fiscal_year = ''
                try:
                    import io
                    raw_bytes = f.read()
                    f.seek(0)
                    if ext.lower() in ('.xlsx',):
                        wb_tmp = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                        ws_tmp = wb_tmp.active
                        for r in range(1, min(6, ws_tmp.max_row + 1)):
                            for c in range(1, min(6, ws_tmp.max_column + 1)):
                                cell_val = str(ws_tmp.cell(r, c).value or '')
                                fy_match = re.search(r'ปีงบประมาณ\s*(\d{4})', cell_val)
                                if fy_match:
                                    fiscal_year = fy_match.group(1)
                                    break
                            if fiscal_year:
                                break
                        wb_tmp.close()
                    elif ext.lower() == '.xls':
                        if xlrd:
                            wb_tmp = xlrd.open_workbook(file_contents=raw_bytes)
                            ws_tmp = wb_tmp.sheet_by_index(0)
                            for r in range(min(5, ws_tmp.nrows)):
                                for c in range(min(5, ws_tmp.ncols)):
                                    cell_val = str(ws_tmp.cell_value(r, c) or '')
                                    fy_match = re.search(r'ปีงบประมาณ\s*(\d{4})', cell_val)
                                    if fy_match:
                                        fiscal_year = fy_match.group(1)
                                        break
                                if fiscal_year:
                                    break
                except Exception:
                    pass

                fy_suffix = f'_ปีงบ{fiscal_year[-2:]}' if fiscal_year else ''
                new_name = f"{prefix}_{branch_name}{fy_suffix}{ext}"

            elif category == 'pending':
                # pending: จะถูกจัดการแบบ batch ด้านล่าง (รวมหลายไฟล์เป็น 1)
                raw_bytes = f.read()
                _pending_batch.append({'filename': filename, 'ext': ext, 'raw_bytes': raw_bytes})
                continue  # ไม่ save ทีละไฟล์ — รอรวมด้านล่าง

            else:
                # fallback ทั่วไป
                m = re.search(r'(\d{6})', name_only)
                if m:
                    new_name = f"{prefix}_{m.group(1)}{ext}"
                else:
                    clean = re.sub(r'[^\w\-.]', '_', name_only).strip('_')[:30]
                    new_name = f"{prefix}_{clean}{ext}"

            dest_path = os.path.join(folder_path, new_name)

            overwritten = os.path.exists(dest_path)
            f.save(dest_path)

            msg = f'{filename} → {new_name}' + (' (เขียนทับ)' if overwritten else '')
            results.append({
                'filename': new_name, 'original': filename,
                'status': 'success', 'message': msg
            })
        except Exception as e:
            errors.append({
                'filename': filename,
                'error': str(e)
            })

    # ── Pending batch merge: รวมหลายไฟล์เป็น 1 ไฟล์ ──
    if category == 'pending' and _pending_batch:
        import io
        try:
            HEADER_ROWS = 8  # Row 0-7 = header area, Row 8+ = data
            DATE_COL = 3     # Col 3 = วันที่แจ้ง (dd/mm/yyyy พ.ศ.)

            all_data_rows = []
            header_rows_cache = None

            for item in _pending_batch:
                raw_bytes = item['raw_bytes']
                ext_lower = item['ext'].lower()

                if ext_lower == '.xls':
                    try:
                        import xlrd
                        wb = xlrd.open_workbook(file_contents=raw_bytes)
                        ws = wb.sheet_by_index(0)
                        # เก็บ header จากไฟล์แรก
                        if header_rows_cache is None:
                            header_rows_cache = []
                            for r in range(min(HEADER_ROWS, ws.nrows)):
                                header_rows_cache.append([ws.cell_value(r, c) for c in range(ws.ncols)])
                        # เก็บ data rows
                        for r in range(HEADER_ROWS, ws.nrows):
                            row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
                            # ข้ามแถวว่าง (ไม่มีเลขแจ้ง)
                            if not str(row_data[2]).strip():
                                continue
                            all_data_rows.append(row_data)
                    except ImportError:
                        errors.append({'filename': item['filename'], 'error': 'xlrd not installed'})

                elif ext_lower == '.xlsx':
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                        ws = wb.active
                        rows_list = list(ws.iter_rows(values_only=True))
                        if header_rows_cache is None:
                            header_rows_cache = [list(r) for r in rows_list[:HEADER_ROWS]]
                        for row in rows_list[HEADER_ROWS:]:
                            row_data = list(row)
                            if not str(row_data[2] if len(row_data) > 2 else '').strip():
                                continue
                            all_data_rows.append(row_data)
                        wb.close()
                    except ImportError:
                        errors.append({'filename': item['filename'], 'error': 'openpyxl not installed'})

            if not all_data_rows:
                errors.append({'filename': 'pending', 'error': 'ไม่พบข้อมูลในไฟล์ที่อัปโหลด'})
            else:
                # ── Parse วันที่แจ้ง เพื่อ sort ──
                def parse_date_for_sort(row):
                    """แปลง dd/mm/yyyy (พ.ศ.) → sortable tuple (yyyy_ce, mm, dd)"""
                    try:
                        val = str(row[DATE_COL]).strip()
                        if '/' in val:
                            parts = val.split('/')
                            dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
                            yyyy_ce = yyyy - 543 if yyyy > 2500 else yyyy
                            return (yyyy_ce, mm, dd)
                    except:
                        pass
                    return (9999, 12, 31)  # ข้อมูลที่ parse ไม่ได้ → ไว้ท้ายสุด

                all_data_rows.sort(key=parse_date_for_sort)

                # ── หาช่วงเดือน (min/max) เพื่อตั้งชื่อ ──
                min_date = (9999, 12)
                max_date = (0, 0)
                for row in all_data_rows:
                    try:
                        val = str(row[DATE_COL]).strip()
                        if '/' in val:
                            parts = val.split('/')
                            mm, yyyy = int(parts[1]), int(parts[2])
                            yy = yyyy - 2500 if yyyy > 2500 else yyyy
                            if (yy, mm) < min_date:
                                min_date = (yy, mm)
                            if (yy, mm) > max_date:
                                max_date = (yy, mm)
                    except:
                        pass

                # ตั้งชื่อ: ค้างซ่อม_MM-YY_to_MM-YY.xlsx
                if min_date[0] < 9999 and max_date[0] > 0:
                    new_name = f"ค้างซ่อม_{min_date[1]:02d}-{min_date[0]:02d}_to_{max_date[1]:02d}-{max_date[0]:02d}.xlsx"
                else:
                    today = datetime.now()
                    yy = today.year - 2543
                    new_name = f"ค้างซ่อม_{today.month:02d}-{yy:02d}.xlsx"

                # ── เขียนไฟล์ .xlsx ──
                import openpyxl
                _ILLEGAL_XML_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = 'ค้างซ่อม'

                def safe_val(v):
                    """ลบ illegal XML characters ที่ .xlsx ไม่รองรับ"""
                    if isinstance(v, str):
                        return _ILLEGAL_XML_RE.sub('', v)
                    return v

                # เขียน header
                if header_rows_cache:
                    for r_idx, row in enumerate(header_rows_cache):
                        for c_idx, val in enumerate(row):
                            ws_out.cell(row=r_idx + 1, column=c_idx + 1, value=safe_val(val))

                # เขียน data (เรียงตามวันที่แจ้งแล้ว)
                start_row = HEADER_ROWS + 1
                for r_idx, row in enumerate(all_data_rows):
                    for c_idx, val in enumerate(row):
                        ws_out.cell(row=start_row + r_idx, column=c_idx + 1, value=safe_val(val))

                dest_path = os.path.join(folder_path, new_name)
                overwritten = os.path.exists(dest_path)
                wb_out.save(dest_path)

                # สร้าง JSON cache ทันทีหลัง upload
                try:
                    _build_json_cache(dest_path)
                except Exception:
                    pass

                orig_names = ', '.join(item['filename'] for item in _pending_batch)
                if overwritten:
                    results.append({
                        'filename': new_name, 'original': orig_names,
                        'status': 'overwrite', 'message': f'เขียนทับ {new_name}'
                    })
                results.append({
                    'filename': new_name, 'original': orig_names,
                    'status': 'success',
                    'message': f'รวม {len(_pending_batch)} ไฟล์ ({len(all_data_rows):,} แถว) → {new_name}'
                })

        except Exception as e:
            errors.append({'filename': 'pending-merge', 'error': str(e)})

    return jsonify({
        'ok': True,
        'category': category,
        'thai_name': CATEGORY_MAP[category],
        'results': results,
        'errors': errors
    })

@app.route('/api/pending-chart')
def api_pending_chart():
    """
    คำนวณข้อมูลกราฟงานค้างซ่อมสะสม (pd1, pd2) จากไฟล์ค้างซ่อม
    นิยาม "ค้างซ่อม ณ สิ้นเดือน M":
      - วันที่แจ้ง <= สิ้นเดือน M
      - AND ( วันเวลาเสร็จสิ้น > สิ้นเดือน M  OR  สถานะ = ซ่อมไม่เสร็จ )
    นับสะสมเฉพาะงานที่แจ้งตั้งแต่ 1 ม.ค. ของปีงบประมาณ
    Query params:
        fy (optional): ปีงบประมาณ พ.ศ. เช่น 2569
    Returns: { ok, fy, fy_list, update_date, months, pd2_data: {month_key: {branch: count}}, pd2_months: [...] }
    """
    import re as _re
    from collections import defaultdict as _ddict
    import calendar

    pending_dir = os.path.join(RAW_DATA_DIR, CATEGORY_MAP['pending'])
    if not os.path.isdir(pending_dir):
        return jsonify({'ok': False, 'error': 'ไม่พบโฟลเดอร์ข้อมูลค้างซ่อม'}), 404

    # Scan files — same logic as pending-table
    fy_files = {}
    for fname in os.listdir(pending_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')):
            continue
        fpath = os.path.join(pending_dir, fname)
        m = _re.search(r'(\d{1,2})-(\d{2})_to_(\d{1,2})-(\d{2})', fname)
        if m:
            start_mm, start_yy = int(m.group(1)), int(m.group(2))
            fy_be = 2500 + start_yy + 1 if start_mm >= 10 else 2500 + start_yy
            fy_files[fy_be] = fpath
        else:
            fy_files.setdefault(0, fpath)

    if not fy_files:
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์ข้อมูลค้างซ่อม'}), 404

    fy_list = sorted([k for k in fy_files.keys() if k > 0])
    req_fy = request.args.get('fy', '')
    fy = int(req_fy) if req_fy and req_fy.isdigit() else (fy_list[-1] if fy_list else 0)
    fpath = fy_files.get(fy, list(fy_files.values())[0])

    if not os.path.isfile(fpath):
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404

    try:
        all_rows = get_pending_rows(fpath)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'อ่านไฟล์ไม่ได้: {e}'}), 500

    # Column indices (0-based)
    col_date = 3        # วันที่แจ้ง (col 4)
    col_finish = 5      # วันเวลาเสร็จสิ้น (col 6)
    col_branch = 19     # สาขา (col 20)
    col_status = 26     # สถานะ (col 27)
    data_start = 8      # row 9 = index 8

    branch_list = [
        "ชลบุรี","พัทยา","บ้านบึง","พนัสนิคม","ศรีราชา","แหลมฉบัง",
        "ฉะเชิงเทรา","บางปะกง","บางคล้า","พนมสารคาม","ระยอง","บ้านฉาง",
        "ปากน้ำประแสร์","จันทบุรี","ขลุง","ตราด","คลองใหญ่",
        "สระแก้ว","วัฒนานคร","อรัญประเทศ","ปราจีนบุรี","กบินทร์บุรี"
    ]

    fy_be = fy if fy > 0 else 2569
    fy_ce = fy_be - 543  # แปลง พ.ศ. → ค.ศ. เพื่อเทียบกับ datetime
    count_start = datetime(fy_ce, 1, 1)

    # Read ALL records into memory
    records = []
    last_report_dt = None
    for row in all_rows[data_start:]:
        if len(row) <= col_status:
            continue
        date_val = row[col_date]
        if not date_val:
            continue
        dt, by = parse_thai_date(date_val)
        if dt is None:
            continue

        if last_report_dt is None or dt > last_report_dt:
            last_report_dt = dt

        finish_val = row[col_finish]
        finish_dt = None
        if finish_val:
            fdt, _ = parse_thai_date(finish_val)
            if fdt:
                finish_dt = fdt
            elif isinstance(finish_val, str):
                try:
                    finish_dt, _ = parse_thai_date(finish_val[:10])
                except:
                    pass

        status = str(row[col_status] or '')
        branch = str(row[col_branch] or '').strip()
        if not branch:
            continue

        records.append({'dt': dt, 'finish_dt': finish_dt, 'status': status, 'branch': branch})

    # Build update_date
    if last_report_dt:
        by_lrd = last_report_dt.year + 543 if last_report_dt.year < 2500 else last_report_dt.year
        update_date = f"{last_report_dt.day:02d}-{last_report_dt.month:02d}-{by_lrd % 100}"
    else:
        update_date = ''

    # Determine which months have data (from Jan of FY year onwards)
    # Find months where we have records from count_start onwards
    month_set = set()
    for rec in records:
        if rec['dt'] >= count_start:
            month_set.add((rec['dt'].year, rec['dt'].month))

    # Sort months and convert to yy-mm format
    sorted_months = sorted(month_set)
    pd2_months = []
    for y, m in sorted_months:
        yy = (y + 543) % 100 if y < 2500 else y % 100
        pd2_months.append(f"{yy:02d}-{m:02d}")

    # For each month, compute "ค้างซ่อม ณ สิ้นเดือน"
    pd2_data = {}  # { "yy-mm": { branch: count } }
    for y, m in sorted_months:
        end_day = calendar.monthrange(y, m)[1]
        end_of_month = datetime(y, m, end_day, 23, 59, 59)
        yy = (y + 543) % 100 if y < 2500 else y % 100
        mk = f"{yy:02d}-{m:02d}"
        branch_counts = _ddict(int)

        for rec in records:
            # Only count jobs reported from count_start up to end_of_month
            if rec['dt'] < count_start or rec['dt'] > end_of_month:
                continue
            # ค้างซ่อม = finish_dt > end_of_month OR status = ซ่อมไม่เสร็จ
            is_pending = False
            if rec['finish_dt'] and rec['finish_dt'] > end_of_month:
                is_pending = True
            elif 'ซ่อมไม่เสร็จ' in rec['status']:
                is_pending = True
            if is_pending:
                branch_counts[rec['branch']] += 1

        pd2_data[mk] = dict(branch_counts)

    # ── Derive pd1_data from pd2_data ──
    # pd1_data[month] = {branch: [prev_month_count, curr_month_count]}
    # prev_month = pd2 snapshot of the previous month (0 if first month)
    pd1_data = {}
    for i, mk in enumerate(pd2_months):
        prev_mk = pd2_months[i - 1] if i > 0 else None
        prev_snap = pd2_data.get(prev_mk, {}) if prev_mk else {}
        curr_snap = pd2_data.get(mk, {})
        branch_pairs = {}
        for b in branch_list:
            pv = prev_snap.get(b, 0)
            cv = curr_snap.get(b, 0)
            branch_pairs[b] = [pv, cv]
        pd1_data[mk] = branch_pairs

    return jsonify({
        'ok': True,
        'fy': fy,
        'fy_list': fy_list,
        'update_date': update_date,
        'pd2_months': pd2_months,
        'pd2_data': pd2_data,
        'pd1_data': pd1_data,
        'branches': branch_list
    })

@app.route('/api/pending-table')
def api_pending_table():
    """
    คำนวณตาราง "งานซ่อมที่ยังไม่ปิดในระบบ" จากไฟล์ค้างซ่อมในโฟลเดอร์ข้อมูลดิบ
    Query params:
        fy (optional): ปีงบประมาณ พ.ศ. เช่น 2569 (default = ล่าสุดที่มี)
    Returns: { ok, fy, fy_list, update_date, branches, months, data: {branch: {month: count}}, col_totals, grand_total }
    """
    import re as _re
    from collections import defaultdict as _ddict

    pending_dir = os.path.join(RAW_DATA_DIR, CATEGORY_MAP['pending'])
    if not os.path.isdir(pending_dir):
        return jsonify({'ok': False, 'error': 'ไม่พบโฟลเดอร์ข้อมูลค้างซ่อม'}), 404

    # Scan all xlsx files — extract fiscal year from filename pattern: ค้างซ่อม_MM-YY_to_MM-YY.xlsx
    fy_files = {}  # { fy_year_be: filepath }
    for fname in os.listdir(pending_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')):
            continue
        fpath = os.path.join(pending_dir, fname)
        # Extract start month from filename: ค้างซ่อม_10-68_to_03-69.xlsx → start=10-68
        m = _re.search(r'(\d{1,2})-(\d{2})_to_(\d{1,2})-(\d{2})', fname)
        if m:
            start_mm, start_yy = int(m.group(1)), int(m.group(2))
            # Fiscal year: if start month is 10 (ต.ค.) → FY = 25(yy+1)
            if start_mm >= 10:
                fy_be = 2500 + start_yy + 1
            else:
                fy_be = 2500 + start_yy
            fy_files[fy_be] = fpath
        else:
            # Fallback: try to read data to determine FY
            fy_files.setdefault(0, fpath)

    if not fy_files:
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์ข้อมูลค้างซ่อม'}), 404

    fy_list = sorted([k for k in fy_files.keys() if k > 0])

    # Determine requested FY
    req_fy = request.args.get('fy', '')
    if req_fy and req_fy.isdigit():
        fy = int(req_fy)
    elif fy_list:
        fy = fy_list[-1]
    else:
        fy = 0

    fpath = fy_files.get(fy, list(fy_files.values())[0])
    if not os.path.isfile(fpath):
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404

    # ── Parse the Excel file (cached) ──
    try:
        all_rows = get_pending_rows(fpath)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'อ่านไฟล์ไม่ได้: {e}'}), 500

    # Column indices (0-based)
    col_date = 3       # วันที่แจ้ง (col 4)
    col_branch = 19    # สาขา (col 20)
    col_status = 26    # สถานะ (col 27)
    data_start = 8     # row 9 = index 8

    # Determine fiscal year months: ต.ค.(YY-1) to ก.ย.(YY)
    fy_yy_start = (fy - 2500 - 1) if fy > 0 else 68
    fy_yy_end = fy_yy_start + 1
    fy_months = []
    for mm in [10, 11, 12]:
        fy_months.append(f"{fy_yy_start:02d}-{mm:02d}")
    for mm in range(1, 10):
        fy_months.append(f"{fy_yy_end:02d}-{mm:02d}")

    # Standard branch list
    branch_list = [
        "ชลบุรี","พัทยา","บ้านบึง","พนัสนิคม","ศรีราชา","แหลมฉบัง",
        "ฉะเชิงเทรา","บางปะกง","บางคล้า","พนมสารคาม","ระยอง","บ้านฉาง",
        "ปากน้ำประแสร์","จันทบุรี","ขลุง","ตราด","คลองใหญ่",
        "สระแก้ว","วัฒนานคร","อรัญประเทศ","ปราจีนบุรี","กบินทร์บุรี"
    ]

    # Count: filter records where สถานะ contains "ซ่อมไม่เสร็จ"
    result = _ddict(lambda: _ddict(int))
    last_report_dt = None

    for row in all_rows[data_start:]:
        if len(row) <= col_status:
            continue
        date_val = row[col_date]
        if not date_val:
            continue

        dt, by = parse_thai_date(date_val)
        if dt is None:
            continue

        if last_report_dt is None or dt > last_report_dt:
            last_report_dt = dt

        status = row[col_status]
        if not status or 'ซ่อมไม่เสร็จ' not in str(status):
            continue

        branch = row[col_branch]
        if not branch:
            continue
        branch = str(branch).strip()

        yy = by % 100
        mm = dt.month
        mk = f"{yy:02d}-{mm:02d}"

        if mk in fy_months:
            result[branch][mk] += 1

    # Build update_date from latest วันที่แจ้ง in data
    if last_report_dt:
        by_lrd = last_report_dt.year + 543 if last_report_dt.year < 2500 else last_report_dt.year
        update_date = f"{last_report_dt.day:02d}-{last_report_dt.month:02d}-{by_lrd % 100}"
    else:
        update_date = ''

    # Build response
    data_out = {}
    col_totals = {mk: 0 for mk in fy_months}
    grand_total = 0
    for branch in branch_list:
        bd = result.get(branch, {})
        if bd:
            data_out[branch] = dict(bd)
            for mk, v in bd.items():
                col_totals[mk] += v
                grand_total += v

    return jsonify({
        'ok': True,
        'fy': fy,
        'fy_be': fy,
        'fy_list': fy_list,
        'update_date': update_date,
        'branches': branch_list,
        'months': fy_months,
        'data': data_out,
        'col_totals': col_totals,
        'grand_total': grand_total
    })

@app.route('/api/pending-detail')
def api_pending_detail():
    """
    รายละเอียดงานค้างซ่อม (สถานะ = ซ่อมไม่เสร็จ) ตามเดือนที่เลือก
    Query params:
        fy: ปีงบประมาณ พ.ศ. (default=2569)
        month: month key เช่น 69-01 (default=all)
        branch: ชื่อสาขา (default=all)
    Returns: { ok, records: [...], total }
    """
    import re as _re

    pending_dir = os.path.join(RAW_DATA_DIR, CATEGORY_MAP['pending'])
    if not os.path.isdir(pending_dir):
        return jsonify({'ok': False, 'error': 'ไม่พบโฟลเดอร์ข้อมูลค้างซ่อม'}), 404

    # Find file by FY (same logic as pending-table)
    fy_files = {}
    for fname in os.listdir(pending_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')) or fname.startswith('~$'):
            continue
        fpath = os.path.join(pending_dir, fname)
        m = _re.search(r'(\d{1,2})-(\d{2})_to_(\d{1,2})-(\d{2})', fname)
        if m:
            start_mm, start_yy = int(m.group(1)), int(m.group(2))
            fy_be = (2500 + start_yy + 1) if start_mm >= 10 else (2500 + start_yy)
            fy_files[fy_be] = fpath
        else:
            fy_files.setdefault(0, fpath)

    if not fy_files:
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์ข้อมูลค้างซ่อม'}), 404

    fy_list = sorted([k for k in fy_files.keys() if k > 0])
    req_fy = request.args.get('fy', '')
    fy = int(req_fy) if req_fy and req_fy.isdigit() else (fy_list[-1] if fy_list else 0)
    fpath = fy_files.get(fy, list(fy_files.values())[0])

    req_month = request.args.get('month', '').strip()
    req_branch = request.args.get('branch', '').strip()

    try:
        rows = get_pending_rows(fpath)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'อ่านไฟล์ไม่ได้: {e}'}), 500

    # Column indices (0-based for cached rows)
    C_DATE = 3       # วันที่แจ้ง (col 4)
    C_NOTIFY = 2     # เลขแจ้ง (col 3)
    C_JOB = 6        # เลขที่งาน (col 7)
    C_TYPE = 7       # ประเภทการร้องเรียน (col 8)
    C_ASPECT = 8     # ด้านการร้องเรียน (col 9)
    C_BRANCH = 19    # สาขา (col 20)
    C_TEAM = 20      # ทีมซ่อม (col 21)
    C_TECH = 21      # ช่างซ่อม (col 22)
    C_PIPE = 25      # ขนาดท่อ (col 26)
    C_STATUS = 26    # สถานะ (col 27)

    data_start = 8  # row 9 = index 8 (0-based)
    records = []

    for row in rows[data_start:]:
        if len(row) <= C_STATUS:
            continue
        status = row[C_STATUS]
        if not status or 'ซ่อมไม่เสร็จ' not in str(status):
            continue

        date_val = row[C_DATE]
        if not date_val:
            continue

        dt, by = parse_thai_date(date_val)
        if dt is None:
            continue

        yy = by % 100
        mk = f"{yy:02d}-{dt.month:02d}"

        # Filter by month if specified
        if req_month and mk != req_month:
            continue

        branch = str(row[C_BRANCH] or '').strip()
        # Filter by branch if specified
        if req_branch and branch != req_branch:
            continue

        rec = {
            'branch': branch,
            'notify_no': str(row[C_NOTIFY] or ''),
            'date': str(date_val) if isinstance(date_val, str) else dt.strftime('%d/%m/%Y'),
            'job_no': str(row[C_JOB] or ''),
            'type': str(row[C_TYPE] or ''),
            'aspect': str(row[C_ASPECT] or ''),
            'team': str(row[C_TEAM] or ''),
            'tech': str(row[C_TECH] or ''),
            'pipe': str(row[C_PIPE] or ''),
            'status': str(row[C_STATUS] or ''),
            'month': mk
        }
        records.append(rec)

    return jsonify({
        'ok': True,
        'records': records,
        'total': len(records)
    })


@app.route('/api/pending-nojob')
def api_pending_nojob():
    """
    ตารางที่ 5: งานที่รับแจ้งท่อแตกรั่วแล้ว แต่ยังไม่เปิดงานซ่อม
    เงื่อนไข:
      - ด้านการร้องเรียน = "ด้านท่อแตกรั่ว" หรือ หัวข้อ/รายละเอียด มีคำว่า ท่อแตก/ท่อรั่ว/แตกรั่ว
      - ยังไม่มีเลขที่งาน (col 6 ว่าง)
      - สถานะ ไม่มีคำว่า "ดำเนินการแล้วเสร็จ"
    Returns: { ok, by_branch: {branch: count}, total, update_date, month_label }
    """
    import re as _re

    pending_dir = os.path.join(RAW_DATA_DIR, CATEGORY_MAP['pending'])
    if not os.path.isdir(pending_dir):
        return jsonify({'ok': False, 'error': 'ไม่พบโฟลเดอร์'}), 404

    fy_files = {}
    for fname in os.listdir(pending_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')) or fname.startswith('~'):
            continue
        fpath = os.path.join(pending_dir, fname)
        m = _re.search(r'(\d{1,2})-(\d{2})_to_(\d{1,2})-(\d{2})', fname)
        if m:
            start_mm, start_yy = int(m.group(1)), int(m.group(2))
            fy_be = 2500 + start_yy + 1 if start_mm >= 10 else 2500 + start_yy
            fy_files[fy_be] = fpath

    if not fy_files:
        return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404

    fy_list = sorted(fy_files.keys())
    req_fy = request.args.get('fy', '')
    fy = int(req_fy) if req_fy and req_fy.isdigit() else (fy_list[-1] if fy_list else 0)
    fpath = fy_files.get(fy, list(fy_files.values())[0])

    try:
        all_rows = get_pending_rows(fpath)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'อ่านไฟล์ไม่ได้: {e}'}), 500

    # Column indices (0-based)
    C_DATE = 3
    C_JOB = 6
    C_SIDE = 8       # ด้านการร้องเรียน
    C_TOPIC = 9      # หัวข้อการร้องเรียน
    C_DETAIL = 10    # รายละเอียดการรับแจ้ง
    C_BRANCH = 19
    C_STATUS = 26
    data_start = 8

    # Find latest month in data
    latest_mk = None
    for row in all_rows[data_start:]:
        if len(row) <= C_BRANCH or not row[C_BRANCH]:
            continue
        date_val = row[C_DATE]
        dt, by = parse_thai_date(date_val)
        if dt and by:
            mm = dt.month
            yy = by % 100
            mk = f"{yy:02d}-{mm:02d}"
            if latest_mk is None or mk > latest_mk:
                latest_mk = mk

    by_branch = {}
    records = []

    for row in all_rows[data_start:]:
        if len(row) <= C_STATUS or not row[C_BRANCH]:
            continue

        branch = str(row[C_BRANCH]).strip()
        date_val = row[C_DATE]
        job_no = row[C_JOB]
        status = str(row[C_STATUS] or '')
        side = str(row[C_SIDE] or '')
        topic = str(row[C_TOPIC] or '')
        detail = str(row[C_DETAIL] or '')

        # Parse date to filter latest month only
        dt, by = parse_thai_date(date_val)
        if not dt or not by:
            continue
        mm = dt.month
        yy = by % 100
        mk = f"{yy:02d}-{mm:02d}"
        if mk != latest_mk:
            continue

        # Condition: pipe complaint
        is_pipe = (side == 'ด้านท่อแตกรั่ว' or
                   'ท่อแตก' in topic or 'ท่อรั่ว' in topic or 'แตกรั่ว' in topic or
                   'ท่อแตก' in detail or 'ท่อรั่ว' in detail or 'แตกรั่ว' in detail)
        if not is_pipe:
            continue

        # Condition: no job number
        has_no_job = (job_no is None or str(job_no).strip() == '')
        if not has_no_job:
            continue

        # Condition: status not done
        if 'ดำเนินการแล้วเสร็จ' in status:
            continue

        by_branch[branch] = by_branch.get(branch, 0) + 1
        records.append({
            'branch': branch,
            'notify_no': str(row[2] or ''),
            'date': str(date_val) if isinstance(date_val, str) else dt.strftime('%d/%m/%Y'),
            'side': side,
            'topic': topic,
            'detail': str(detail)[:100],
            'status': status
        })

    # Get update date from file mod time
    import time as _time
    fmod = os.path.getmtime(fpath)
    update_date = _time.strftime('%d/%m/%Y', _time.localtime(fmod))

    # Month label
    month_label = latest_mk or ''

    return jsonify({
        'ok': True,
        'by_branch': by_branch,
        'records': records,
        'total': len(records),
        'update_date': update_date,
        'month_key': latest_mk,
        'fy': fy
    })


@app.route('/api/data/<category>/<filename>', methods=['DELETE', 'OPTIONS'])
def api_delete_file(category, filename):
    """ลบไฟล์เฉพาะ"""
    if request.method == 'OPTIONS':
        return '', 204

    if category not in CATEGORY_MAP:
        return jsonify({'ok': False, 'error': f'ไม่รู้จัก category: {category}'}), 400

    folder_path = os.path.join(RAW_DATA_DIR, CATEGORY_MAP[category])
    file_path = os.path.join(folder_path, filename)

    # Safety check: ensure file is within folder
    if not os.path.abspath(file_path).startswith(os.path.abspath(folder_path)):
        return jsonify({'ok': False, 'error': 'ไม่อนุญาต'}), 403

    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            return jsonify({'ok': True, 'filename': filename, 'deleted': True})
        else:
            return jsonify({'ok': False, 'error': 'ไม่พบไฟล์'}), 404
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/open-folder', methods=['POST', 'OPTIONS'])
def api_open_folder():
    """เปิดโฟลเดอร์ ข้อมูลดิบ/ ใน File Explorer"""
    if request.method == 'OPTIONS':
        return '', 204

    try:
        if platform.system() == 'Windows':
            os.startfile(RAW_DATA_DIR)
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', RAW_DATA_DIR])
        else:
            subprocess.Popen(['xdg-open', RAW_DATA_DIR])

        return jsonify({'ok': True, 'path': RAW_DATA_DIR})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/open-main', methods=['POST', 'OPTIONS'])
def api_open_main():
    """เปิดหน้าหลัก (Landing Page) ใน browser"""
    if request.method == 'OPTIONS':
        return '', 204

    try:
        parent_dir = os.path.dirname(BASE_DIR)
        main_file = os.path.join(parent_dir, 'index.html')

        if platform.system() == 'Windows':
            os.startfile(main_file)
        else:
            subprocess.Popen(['xdg-open', main_file])

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/rebuild', methods=['POST', 'OPTIONS'])
def api_rebuild():
    """รัน build_dashboard.py เพื่อสร้าง dashboard ใหม่"""
    if request.method == 'OPTIONS':
        return '', 204

    try:
        build_script = os.path.join(BASE_DIR, 'build_dashboard.py')

        if not os.path.exists(build_script):
            return jsonify({'ok': False, 'error': 'ไม่พบไฟล์ build_dashboard.py'}), 404

        # Run build script
        result = subprocess.run(
            [sys.executable, build_script],
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
            timeout=120
        )

        return jsonify({
            'ok': result.returncode == 0,
            'message': 'สร้าง Dashboard สำเร็จ' if result.returncode == 0 else 'เกิดข้อผิดพลาด',
            'stdout': result.stdout[-500:] if result.stdout else '',  # Last 500 chars
            'stderr': result.stderr[-500:] if result.stderr else '',
            'returncode': result.returncode
        })
    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Timeout — build script ใช้เวลานานเกินไป'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ─── Catch-all Static Files (must be AFTER all API routes) ────────────────────

@app.route('/<path:path>')
def serve_static(path):
    full_path = os.path.join(BASE_DIR, path)
    if os.path.isfile(full_path):
        return send_from_directory(BASE_DIR, path)
    return send_from_directory(BASE_DIR, 'index.html')

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("=" * 60)
    print("  Dashboard GIS — Development Server")
    print(f"  http://localhost:{PORT}")
    print("=" * 60)
    print(f"  Base directory : {BASE_DIR}")
    print(f"  Raw data dir   : {RAW_DATA_DIR}")
    print()

    # Pre-build JSON cache สำหรับไฟล์ที่มีอยู่แล้ว (ทำครั้งเดียวตอน start)
    preload_pending_cache()
    print()

    app.run(host='0.0.0.0', port=PORT, debug=True, threaded=True)
