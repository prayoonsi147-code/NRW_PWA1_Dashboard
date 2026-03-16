#!/usr/bin/env python3
"""
build_dashboard.py - สร้าง Dashboard งานลูกค้าสัมพันธ์ กปภ.เขต 1
อ่านข้อมูลจาก Excel แล้ว embed ลงใน index.html
"""
import openpyxl
import json
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "ข้อมูลดิบ", "เรื่องร้องเรียน")
HTML_TEMPLATE = os.path.join(SCRIPT_DIR, "index.html")

def clean_num(val):
    """Clean number value from Excel"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    s = str(val).replace(',', '').replace('\xa0', '').replace(' ', '').strip()
    if s == '' or s == '-':
        return 0
    try:
        return float(s)
    except:
        return 0

def read_all_data():
    """Read all Excel files and return structured data"""
    cat_names = [
        'ด้านปริมาณน้ำ', 'ด้านท่อแตกรั่ว', 'ด้านคุณภาพน้ำ',
        'ด้านการบริการ', 'ด้านบุคลากร', 'การแจ้งเหตุ',
        'ด้านการติดตามเร่งรัดข้อร้องเรียนเดิม', 'ด้านสอบถามทั่วไป',
        'ความต้องการ ความคาดหวัง และข้อเสนอแนะ', 'คำชม และอื่นๆ'
    ]

    files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')])
    print(f"  พบไฟล์ข้อมูล: {len(files)} ไฟล์")

    all_data = {}
    branches_order = []

    for fname in files:
        match = re.search(r'(\d{2})-(\d{2})', fname)
        if not match:
            continue
        year_be = int(match.group(1))
        month = int(match.group(2))
        month_key = f"{year_be:02d}-{month:02d}"

        try:
            wb = openpyxl.load_workbook(os.path.join(DATA_DIR, fname), data_only=True)
        except Exception as e:
            print(f"  [WARNING] ข้ามไฟล์เสีย: {fname} ({e})")
            continue
        ws = wb[wb.sheetnames[0]]
        month_data = {}

        for row_idx in range(7, 29):
            branch = ws.cell(row=row_idx, column=2).value
            if not branch:
                continue

            customers = clean_num(ws.cell(row=row_idx, column=3).value)
            categories = {}
            col_start = 5
            for i, cat_name in enumerate(cat_names):
                col = col_start + i * 3
                categories[cat_name] = {
                    'รวม': clean_num(ws.cell(row=row_idx, column=col).value),
                    'ไม่เกิน': clean_num(ws.cell(row=row_idx, column=col+1).value),
                    'เกิน': clean_num(ws.cell(row=row_idx, column=col+2).value)
                }
            total = clean_num(ws.cell(row=row_idx, column=35).value)
            total_w = clean_num(ws.cell(row=row_idx, column=36).value)
            total_o = clean_num(ws.cell(row=row_idx, column=37).value)

            month_data[branch] = {
                'จำนวนลูกค้า': customers,
                'categories': categories,
                'รวมสาขา': total,
                'รวม_ไม่เกิน': total_w,
                'รวม_เกิน': total_o
            }
            if branch not in branches_order:
                branches_order.append(branch)

        # Row 29: Region 1 total
        row_idx = 29
        customers = clean_num(ws.cell(row=row_idx, column=3).value)
        categories = {}
        col_start = 5
        for i, cat_name in enumerate(cat_names):
            col = col_start + i * 3
            categories[cat_name] = {
                'รวม': clean_num(ws.cell(row=row_idx, column=col).value),
                'ไม่เกิน': clean_num(ws.cell(row=row_idx, column=col+1).value),
                'เกิน': clean_num(ws.cell(row=row_idx, column=col+2).value)
            }
        total = clean_num(ws.cell(row=row_idx, column=35).value)
        total_w = clean_num(ws.cell(row=row_idx, column=36).value)
        total_o = clean_num(ws.cell(row=row_idx, column=37).value)
        month_data['รวม เขต 1'] = {
            'จำนวนลูกค้า': customers,
            'categories': categories,
            'รวมสาขา': total,
            'รวม_ไม่เกิน': total_w,
            'รวม_เกิน': total_o
        }

        all_data[month_key] = month_data
        wb.close()

    # Determine 13-month range
    months_sorted = sorted(all_data.keys())
    latest = months_sorted[-1]
    ly, lm = int(latest[:2]), int(latest[3:])
    same_month_ly = f"{ly-1:02d}-{lm:02d}"
    months_13 = [m for m in months_sorted if m >= same_month_ly]

    print(f"  ช่วงเดือน: {months_13[0]} - {months_13[-1]} ({len(months_13)} เดือน)")
    print(f"  จำนวนสาขา: {len(branches_order)}")

    return {
        'months': months_13,
        'branches': branches_order,
        'all_months': months_sorted,
        'data': all_data,
        'cat_names': cat_names
    }

def build():
    print("=" * 50)
    print("  Build Dashboard งานลูกค้าสัมพันธ์ กปภ.เขต 1")
    print("=" * 50)

    print("\n[1/3] อ่านข้อมูล Excel...")
    data = read_all_data()

    print("\n[2/3] สร้าง JSON...")
    data_json = json.dumps(data, ensure_ascii=False)
    print(f"  ขนาดข้อมูล: {len(data_json):,} bytes")

    print("\n[3/3] Embed ข้อมูลลงใน index.html...")
    with open(HTML_TEMPLATE, 'r', encoding='utf-8') as f:
        html = f.read()

    # Try placeholder first (fresh template)
    if 'DASHBOARD_DATA_PLACEHOLDER' in html:
        html = html.replace('DASHBOARD_DATA_PLACEHOLDER', data_json)
        print("  (ใช้ placeholder)")
    else:
        # Replace existing embedded DATA (previously built file)
        # DATA is on a single line: const DATA = {...};
        pattern = r'^const DATA = \{.*\};$'
        new_val = 'const DATA = ' + data_json + ';'
        html_new, count = re.subn(pattern, new_val, html, count=1, flags=re.MULTILINE)
        if count > 0:
            html = html_new
            print("  (แทนที่ const DATA เดิม)")
        else:
            print("  [WARNING] ไม่พบ DASHBOARD_DATA_PLACEHOLDER หรือ const DATA ใน HTML!")

    output_path = os.path.join(SCRIPT_DIR, "index.html")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(output_path) / 1024
    print(f"  บันทึก: {output_path}")
    print(f"  ขนาดไฟล์: {size_kb:.1f} KB")
    print("\n  เสร็จสิ้น!")

if __name__ == '__main__':
    build()
