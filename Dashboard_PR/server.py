# -*- coding: utf-8 -*-
"""
Dashboard PR — Local Development Server
========================================
Flask server สำหรับ Dashboard ข้อร้องเรียน + Always-On
รับ upload ไฟล์ Excel → auto-detect เดือน → auto-rename → parse → เก็บถาวร

Usage:
    python server.py
    หรือ ดับเบิลคลิก start_server.bat

API Endpoints:
    GET  /                     → serve index.html
    GET  /api/data             → ข้อมูลทั้งหมดที่ upload ไว้ (JSON)
    POST /api/upload/pr        → upload ไฟล์ PR (GUI_019)
    POST /api/upload/aon       → upload ไฟล์ AON (Always-On)
    DELETE /api/data/pr/<mk>   → ลบข้อมูล PR เดือนที่ระบุ
    DELETE /api/data/aon/<mk>  → ลบข้อมูล AON เดือนที่ระบุ
    POST /api/data/edit/pr     → แก้ไขข้อมูล PR
    POST /api/data/edit/aon    → แก้ไขข้อมูล AON
    GET  /api/files            → รายชื่อไฟล์ที่ upload ไว้
    DELETE /api/data/clear     → ล้างข้อมูลทั้งหมด
"""

from flask import Flask, request, jsonify, send_from_directory
import os, sys, json, re, shutil, traceback
from datetime import datetime

# --- Excel Libraries ---
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DATA_DIR = os.path.join(BASE_DIR, 'ข้อมูลดิบ')
UPLOAD_DIR = RAW_DATA_DIR  # backward compat alias
PR_DIR = os.path.join(RAW_DATA_DIR, 'เรื่องร้องเรียน')
AON_DIR = os.path.join(RAW_DATA_DIR, 'AlwayON')
DATA_FILE = os.path.join(RAW_DATA_DIR, 'data.json')
PORT = 5000

# Create directories
os.makedirs(PR_DIR, exist_ok=True)
os.makedirs(AON_DIR, exist_ok=True)

# ─── Flask App ────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder=BASE_DIR)

# ─── CORS Middleware ──────────────────────────────────────────────────────────

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

PARENT_DIR = os.path.dirname(BASE_DIR)

@app.route('/api/open-main', methods=['POST', 'OPTIONS'])
def api_open_main():
    """เปิดหน้าหลัก (Landing Page) ใน browser"""
    if request.method == 'OPTIONS':
        return '', 204
    import platform
    main_file = os.path.join(PARENT_DIR, 'index.html')
    try:
        if platform.system() == 'Windows':
            os.startfile(main_file)
        else:
            import subprocess
            subprocess.Popen(['xdg-open', main_file])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_static(path):
    if not path or path == '':
        return send_from_directory(BASE_DIR, 'index.html')
    full_path = os.path.join(BASE_DIR, path)
    if os.path.isfile(full_path):
        return send_from_directory(BASE_DIR, path)
    return send_from_directory(BASE_DIR, 'index.html')

# ─── Branch Name Normalization ────────────────────────────────────────────────

BRANCH_NAME_MAP = {
    'สาขาชลบุรี': 'ชลบุรี',
    'สาขาบ้านบึง': 'บ้านบึง',
    'สาขาพนัสนิคม': 'พนัสนิคม',
    'สาขาศรีราชา': 'ศรีราชา',
    'สาขาแหลมฉบัง': 'แหลมฉบัง',
    'สาขาพัทยา': 'พัทยา',
    'สาขาฉะเชิงเทรา': 'ฉะเชิงเทรา',
    'สาขาบางปะกง': 'บางปะกง',
    'สาขาบางคล้า': 'บางคล้า',
    'สาขาพนมสารคาม': 'พนมสารคาม',
    'สาขาระยอง': 'ระยอง',
    'สาขาบ้านฉาง': 'บ้านฉาง',
    'สาขาปากน้ำประแสร์': 'ปากน้ำประแสร์',
    'สาขาจันทบุรี': 'จันทบุรี',
    'สาขาขลุง': 'ขลุง',
    'สาขาตราด': 'ตราด',
    'สาขาคลองใหญ่': 'คลองใหญ่',
    'สาขาสระแก้ว': 'สระแก้ว',
    'สาขาวัฒนานคร': 'วัฒนานคร',
    'สาขาอรัญประเทศ': 'อรัญประเทศ',
    'สาขาปราจีนบุรี': 'ปราจีนบุรี',
    'สาขากบินทร์บุรี': 'กบินทร์บุรี',
    # พิเศษ
    'ชลบุรี(พ)': 'ชลบุรี',
    'พัทยา(พ)': 'พัทยา',
    'ชลบุรี(พิเศษ)': 'ชลบุรี',
    'พัทยา(พิเศษ)': 'พัทยา',
}

def norm_branch(name):
    """Normalize branch name — ตัด 'สาขา' นำหน้า + map ชื่อพิเศษ"""
    n = str(name).strip()
    if n in BRANCH_NAME_MAP:
        return BRANCH_NAME_MAP[n]
    if n.startswith('สาขา'):
        return n[len('สาขา'):]
    return n

# ─── Thai Month Maps ─────────────────────────────────────────────────────────

THAI_MONTH_ABBR = {
    'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4, 'พ.ค.': 5, 'มิ.ย.': 6,
    'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12,
}

THAI_MONTH_FULL = {
    'มกราคม': 1, 'กุมภาพันธ์': 2, 'มีนาคม': 3, 'เมษายน': 4,
    'พฤษภาคม': 5, 'มิถุนายน': 6, 'กรกฎาคม': 7, 'สิงหาคม': 8,
    'กันยายน': 9, 'ตุลาคม': 10, 'พฤศจิกายน': 11, 'ธันวาคม': 12,
}

def make_month_key(yyyy, mm):
    """Convert พ.ศ. year + month to 'YY-MM' format"""
    yy = yyyy - 2500
    return f"{yy:02d}-{mm:02d}"

# ─── Month Detection ─────────────────────────────────────────────────────────

def detect_month_from_filename(filename):
    """ตรวจจับเดือนจากชื่อไฟล์ — หา pattern YY-MM"""
    m = re.search(r'(\d{2})-(\d{2})', filename)
    if m:
        yy, mm = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12:
            return f"{yy:02d}-{mm:02d}"
    return None

def detect_month_from_text(text):
    """ตรวจจับเดือนจากข้อความ เช่น 'ถึง เดือน มกราคม พ.ศ. 2569'"""
    if not text:
        return None
    s = str(text)
    month_num = None
    year_num = None

    # หาชื่อเดือนเต็ม
    for name, num in THAI_MONTH_FULL.items():
        if name in s:
            month_num = num
            break
    # หาชื่อเดือนย่อ
    if month_num is None:
        for abbr, num in THAI_MONTH_ABBR.items():
            if abbr in s:
                month_num = num
                break

    # หาปี (4 หลัก)
    m = re.search(r'(\d{4})', s)
    if m:
        year_num = int(m.group(1))
    # หาปี (2 หลัก) ถ้าไม่เจอ 4 หลัก
    if year_num is None:
        m2 = re.search(r'(\d{2})', s)
        if m2:
            yy = int(m2.group(1))
            year_num = yy + 2500

    if month_num and year_num:
        return make_month_key(year_num, month_num)
    return None

def sheet_name_to_month_key(name):
    """แปลงชื่อ Sheet (เช่น 'ต.ค.68') เป็น month key"""
    s = str(name).strip()
    for abbr, mm in THAI_MONTH_ABBR.items():
        if abbr in s:
            m = re.search(r'(\d{2})', s)
            if m:
                yy = int(m.group(1))
                return f"{yy:02d}-{mm:02d}"
    return None

def parse_always_on_header(header):
    """ตรวจ header ที่มีคำว่า 'always on' + ตรวจจับเดือน"""
    s = str(header).strip().lower()
    if 'always on' not in s and 'always-on' not in s:
        return None
    return detect_month_from_text(str(header))

# ─── Excel Reading Helpers ────────────────────────────────────────────────────

def read_excel_rows(filepath):
    """อ่านไฟล์ Excel (.xlsx/.xls) คืน list of sheets, แต่ละ sheet = {name, rows}"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xlsx' or ext == '.xlsm':
        if openpyxl is None:
            raise ImportError("ไม่พบ library openpyxl — กรุณาติดตั้ง: pip install openpyxl")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            sheets.append({'name': name, 'rows': rows})
        wb.close()
        return sheets

    elif ext == '.xls':
        if xlrd is None:
            raise ImportError("ไม่พบ library xlrd — กรุณาติดตั้ง: pip install xlrd")
        wb = xlrd.open_workbook(filepath)
        sheets = []
        for idx in range(wb.nsheets):
            ws = wb.sheet_by_index(idx)
            rows = []
            for r in range(ws.nrows):
                rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
            sheets.append({'name': ws.name, 'rows': rows})
        return sheets

    else:
        raise ValueError(f"ไม่รองรับนามสกุลไฟล์ {ext} (รองรับ .xlsx, .xlsm, .xls)")

def parse_num(val):
    """แปลงค่าเป็นตัวเลข — คืน 0 ถ้าแปลงไม่ได้"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val if val == val else 0  # NaN check
    try:
        s = str(val).replace(',', '').strip()
        return float(s) if s else 0
    except (ValueError, TypeError):
        return 0

# ─── PR Parser (GUI_019 Format) ──────────────────────────────────────────────

def parse_pr_file(filepath, filename, manual_mk=None):
    """
    Parse ไฟล์ PR (GUI_019 format)
    Returns: {mk: str, data: {branch: {...}}, count: int, cat_names: [str]}
    """
    sheets = read_excel_rows(filepath)
    if not sheets:
        raise ValueError(f"[{filename}] ไฟล์ว่างเปล่า")

    rows = sheets[0]['rows']
    if len(rows) < 7:
        raise ValueError(f"[{filename}] ไฟล์มีข้อมูลไม่เพียงพอ ({len(rows)} แถว)")

    # --- Detect month ---
    mk = manual_mk
    if not mk:
        mk = detect_month_from_filename(filename)
    if not mk and len(rows) > 2:
        for cell in rows[2]:
            mk = detect_month_from_text(cell)
            if mk:
                break
    if not mk:
        raise ValueError(f"[{filename}] ไม่สามารถตรวจจับเดือนได้")

    # --- Parse header row (row index 4) for categories ---
    header_row = rows[4] if len(rows) > 4 else []
    cat_groups = []
    c = 4
    while c < len(header_row):
        h = str(header_row[c] or '').strip()
        if h and h != 'รวมสาขา':
            cat_name = re.sub(r'^\d+\.\s*', '', h).strip()
            cat_groups.append({'name': cat_name, 'start_col': c})
            c += 3  # each category = 3 columns (รวม, ไม่เกิน, เกิน)
        elif h == 'รวมสาขา':
            break
        else:
            c += 1

    # Find รวมสาขา column
    total_col = -1
    for c2 in range(4, len(header_row)):
        if str(header_row[c2] or '').strip() == 'รวมสาขา':
            total_col = c2
            break

    # --- Parse data rows (row index 6+) ---
    branch_data = {}
    cat_names = [cg['name'] for cg in cat_groups]

    for r in range(6, len(rows)):
        row = rows[r]
        if not row or len(row) < 5:
            continue
        branch_name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not branch_name:
            continue
        # เก็บ "รวม เขต 1" ไว้ (ใช้แสดงภาพรวมในกราฟ) แต่ข้าม "รวมทั้งหมด" / "รวม"
        is_regional = branch_name in ('รวม เขต 1',)
        if branch_name in ('รวมทั้งหมด', 'รวม'):
            continue
        if not is_regional:
            branch_name = norm_branch(branch_name)
        if not branch_name:
            continue

        bd = {}
        bd['จำนวนลูกค้า'] = parse_num(row[2] if len(row) > 2 else 0)
        bd['categories'] = {}

        for cg in cat_groups:
            sc = cg['start_col']
            bd['categories'][cg['name']] = {
                'รวม': parse_num(row[sc] if len(row) > sc else 0),
                'ไม่เกิน': parse_num(row[sc + 1] if len(row) > sc + 1 else 0),
                'เกิน': parse_num(row[sc + 2] if len(row) > sc + 2 else 0),
            }

        # รวมสาขา
        if total_col >= 0 and len(row) > total_col:
            bd['รวมสาขา'] = parse_num(row[total_col])
            bd['รวม_ไม่เกิน'] = parse_num(row[total_col + 1] if len(row) > total_col + 1 else 0)
            bd['รวม_เกิน'] = parse_num(row[total_col + 2] if len(row) > total_col + 2 else 0)
        else:
            tot = sum(parse_num(row[cg['start_col']] if len(row) > cg['start_col'] else 0) for cg in cat_groups)
            tot_ne = sum(parse_num(row[cg['start_col'] + 1] if len(row) > cg['start_col'] + 1 else 0) for cg in cat_groups)
            tot_e = sum(parse_num(row[cg['start_col'] + 2] if len(row) > cg['start_col'] + 2 else 0) for cg in cat_groups)
            bd['รวมสาขา'] = tot
            bd['รวม_ไม่เกิน'] = tot_ne
            bd['รวม_เกิน'] = tot_e

        branch_data[branch_name] = bd

    return {
        'mk': mk,
        'data': branch_data,
        'count': len(branch_data),
        'cat_names': cat_names,
    }

# ─── AON Parser (Always-On) ──────────────────────────────────────────────────

def parse_aon_sheet_with_col(rows, aon_col, month_key):
    """Parse AON data จาก sheet ที่ระบุ column"""
    result = {}

    # หา column ชื่อหน่วยงาน (row index 3)
    name_col = 4
    dist_col = -1
    if len(rows) > 3:
        for c, val in enumerate(rows[3]):
            h = str(val or '').strip()
            if h == 'หน่วยงาน':
                name_col = c
            elif h == 'เขต':
                dist_col = c

    # อ่านข้อมูล (row index 5+)
    for r in range(5, len(rows)):
        row = rows[r]
        if not row or len(row) <= aon_col:
            continue

        # Filter เฉพาะ เขต 1
        if dist_col >= 0:
            dist = parse_num(row[dist_col] if len(row) > dist_col else 0)
            if dist != 1:
                continue

        raw_name = str(row[name_col] or '').strip() if len(row) > name_col else ''
        if not raw_name:
            continue
        branch_name = norm_branch(raw_name)
        if not branch_name or branch_name.startswith('รวม'):
            continue

        val = row[aon_col]
        if val is None or val == '':
            continue
        try:
            num_val = float(val)
        except (ValueError, TypeError):
            continue

        # ค่า 0-1 → คูณ 100 เป็น %
        if num_val <= 1.5:
            num_val = round(num_val * 10000) / 100

        result[branch_name] = num_val

    return result

def parse_aon_file(filepath, filename, mode='auto', manual_mk=None):
    """
    Parse ไฟล์ AON (Always-On)
    Returns: {months: {mk: {branch: value}}, count: int, processed_months: [str]}
    """
    sheets = read_excel_rows(filepath)
    if not sheets:
        raise ValueError(f"[{filename}] ไฟล์ว่างเปล่า")

    all_data = {}
    total_count = 0
    processed_months = []

    if mode == 'manual' and manual_mk:
        # Manual mode: อ่าน sheet แรก ใช้ manual_mk
        rows = sheets[0]['rows']
        if len(rows) < 6:
            raise ValueError(f"[{filename}] ข้อมูลไม่เพียงพอ")
        # หา always on column
        sub_header = rows[4] if len(rows) > 4 else []
        aon_col = -1
        for c, val in enumerate(sub_header):
            if parse_always_on_header(str(val or '')):
                aon_col = c
                break
        if aon_col < 0:
            # ถ้าหาไม่เจอ ลองใช้ column สุดท้ายที่มีค่า
            raise ValueError(f"[{filename}] ไม่พบคอลัมน์ always on")
        data = parse_aon_sheet_with_col(rows, aon_col, manual_mk)
        if data:
            all_data[manual_mk] = data
            total_count += len(data)
            processed_months.append(manual_mk)
    else:
        # Auto mode: scan ทุก sheet
        for sheet in sheets:
            rows = sheet['rows']
            if len(rows) < 6:
                continue

            sub_header = rows[4] if len(rows) > 4 else []
            aon_cols = []
            for c, val in enumerate(sub_header):
                mk2 = parse_always_on_header(str(val or ''))
                if mk2:
                    aon_cols.append({'col': c, 'mk': mk2})

            if not aon_cols:
                continue

            # เลือก column ที่ match กับชื่อ sheet (ถ้าได้)
            sheet_mk = sheet_name_to_month_key(sheet['name'])
            best_col = None
            for ac in aon_cols:
                if ac['mk'] == sheet_mk:
                    best_col = ac
                    break
            if not best_col:
                best_col = aon_cols[0]

            data = parse_aon_sheet_with_col(rows, best_col['col'], best_col['mk'])
            if data:
                all_data[best_col['mk']] = data
                total_count += len(data)
                if best_col['mk'] not in processed_months:
                    processed_months.append(best_col['mk'])

    return {
        'months': all_data,
        'count': total_count,
        'processed_months': sorted(processed_months),
    }

# ─── Data Persistence (JSON) ─────────────────────────────────────────────────

def load_data():
    """โหลดข้อมูลจาก data.json"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {'pr': {}, 'aon': {}, 'pr_files': {}, 'aon_files': {}, 'pr_cat_names': []}

def save_data(data):
    """บันทึกข้อมูลลง data.json"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ─── Excel Write-Back ─────────────────────────────────────────────────────────

def write_back_pr_excel(filepath, edit_data, cat_names):
    """
    เขียนข้อมูลที่แก้ไขกลับลงไฟล์ PR Excel (.xlsx)
    edit_data = {branch: {จำนวนลูกค้า, categories: {cat: {รวม, ไม่เกิน, เกิน}}, รวมสาขา, ...}}
    """
    if openpyxl is None:
        return False

    ext = os.path.splitext(filepath)[1].lower()
    if ext != '.xlsx' and ext != '.xlsm':
        return False

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # --- Parse header to find column mapping (same logic as parse_pr_file) ---
    header_row = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
    cat_col_map = {}  # cat_name → start_col (1-indexed)
    c = 5  # start from column E (1-indexed = 5)
    while c <= len(header_row):
        h = str(header_row[c - 1] or '').strip()
        if h and h != 'รวมสาขา':
            cat_name = re.sub(r'^\d+\.\s*', '', h).strip()
            cat_col_map[cat_name] = c
            c += 3
        elif h == 'รวมสาขา':
            break
        else:
            c += 1

    # Find รวมสาขา column
    total_col = -1
    for c2 in range(5, len(header_row) + 1):
        if str(header_row[c2 - 1] or '').strip() == 'รวมสาขา':
            total_col = c2
            break

    # --- Build branch → row mapping (row 7+) ---
    branch_row_map = {}
    for r in range(7, ws.max_row + 1):
        raw_name = str(ws.cell(row=r, column=2).value or '').strip()
        if not raw_name:
            continue
        if raw_name in ('รวม เขต 1', 'รวมทั้งหมด', 'รวม'):
            continue
        bn = norm_branch(raw_name)
        if bn:
            branch_row_map[bn] = r

    # --- Write data back ---
    updated = False
    for branch, bd in edit_data.items():
        row_num = branch_row_map.get(branch)
        if not row_num:
            continue

        # จำนวนลูกค้า (column C = 3)
        if 'จำนวนลูกค้า' in bd:
            ws.cell(row=row_num, column=3, value=bd['จำนวนลูกค้า'])

        # Categories
        cats = bd.get('categories', {})
        for cat_name, vals in cats.items():
            sc = cat_col_map.get(cat_name)
            if not sc:
                continue
            ws.cell(row=row_num, column=sc, value=vals.get('รวม', 0))
            ws.cell(row=row_num, column=sc + 1, value=vals.get('ไม่เกิน', 0))
            ws.cell(row=row_num, column=sc + 2, value=vals.get('เกิน', 0))

        # รวมสาขา
        if total_col > 0:
            ws.cell(row=row_num, column=total_col, value=bd.get('รวมสาขา', 0))
            ws.cell(row=row_num, column=total_col + 1, value=bd.get('รวม_ไม่เกิน', 0))
            ws.cell(row=row_num, column=total_col + 2, value=bd.get('รวม_เกิน', 0))

        updated = True

    if updated:
        wb.save(filepath)
    wb.close()
    return updated


def write_back_aon_excel(filepath, mk, edit_data):
    """
    เขียนข้อมูลที่แก้ไขกลับลงไฟล์ AON Excel (.xls → convert to .xlsx / .xlsx direct)
    edit_data = {branch: value (percentage)}
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xlsx' and openpyxl is not None:
        return _write_back_aon_xlsx(filepath, mk, edit_data)
    elif ext == '.xls' and openpyxl is not None:
        # .xls (BIFF8) → ต้องแปลงเป็น .xlsx ก่อนถึงเขียนได้
        # อ่าน .xls ด้วย xlrd → เขียนใหม่เป็น .xlsx ด้วย openpyxl
        return _convert_and_write_back_aon_xls(filepath, mk, edit_data)
    return False


def _write_back_aon_xlsx(filepath, mk, edit_data):
    """เขียนกลับ AON .xlsx"""
    wb = openpyxl.load_workbook(filepath)

    # หา sheet + column ที่ตรงกับ mk
    for ws in wb.worksheets:
        # ลองหา always on column ใน sub-header (row 5)
        sub_header = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column + 1)]
        for c_idx, val in enumerate(sub_header):
            col_mk = parse_always_on_header(str(val or ''))
            if col_mk == mk:
                aon_col = c_idx + 1  # 1-indexed

                # หา column ชื่อหน่วยงาน (row 4)
                name_col = 5  # default
                for nc in range(1, ws.max_column + 1):
                    if str(ws.cell(row=4, column=nc).value or '').strip() == 'หน่วยงาน':
                        name_col = nc
                        break

                # เขียนกลับ
                updated = False
                for r in range(6, ws.max_row + 1):
                    raw_name = str(ws.cell(row=r, column=name_col).value or '').strip()
                    if not raw_name:
                        continue
                    bn = norm_branch(raw_name)
                    if bn and bn in edit_data:
                        val_pct = edit_data[bn]
                        # เขียนกลับเป็น 0-1 scale ถ้าค่าเดิมเป็น 0-1
                        original = ws.cell(row=r, column=aon_col).value
                        if original is not None and isinstance(original, (int, float)) and original <= 1.5:
                            ws.cell(row=r, column=aon_col, value=round(val_pct / 100, 6))
                        else:
                            ws.cell(row=r, column=aon_col, value=val_pct)
                        updated = True

                if updated:
                    wb.save(filepath)
                    wb.close()
                    return True

    wb.close()
    return False


def _convert_and_write_back_aon_xls(filepath, mk, edit_data):
    """
    .xls (BIFF8) → อ่านด้วย xlrd → เขียนใหม่ .xlsx ด้วย openpyxl พร้อมข้อมูลที่แก้ไข
    เปลี่ยนนามสกุลไฟล์เป็น .xlsx
    """
    if xlrd is None or openpyxl is None:
        return False

    # อ่าน .xls ทั้งหมด
    rb = xlrd.open_workbook(filepath)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    target_sheet_idx = -1
    target_aon_col = -1
    target_name_col = 4  # 0-indexed default

    for si in range(rb.nsheets):
        rs = rb.sheet_by_index(si)
        ws_new = wb_new.create_sheet(title=rs.name)

        # Copy all cells
        for r in range(rs.nrows):
            for c in range(rs.ncols):
                ws_new.cell(row=r + 1, column=c + 1, value=rs.cell_value(r, c))

        # ลองหา always on column ใน row index 4 (sub-header)
        if rs.nrows > 4:
            for c in range(rs.ncols):
                val = str(rs.cell_value(4, c) or '').strip()
                col_mk = parse_always_on_header(val)
                if col_mk == mk:
                    target_sheet_idx = si
                    target_aon_col = c

            # หา column ชื่อหน่วยงาน (row index 3)
            if target_sheet_idx == si:
                for c in range(rs.ncols):
                    if str(rs.cell_value(3, c) or '').strip() == 'หน่วยงาน':
                        target_name_col = c
                        break

    rb.release_resources()

    # แก้ไขข้อมูลใน sheet ที่ตรง
    if target_sheet_idx >= 0 and target_aon_col >= 0:
        ws_target = wb_new.worksheets[target_sheet_idx]
        updated = False

        for r in range(6, ws_target.max_row + 1):
            raw_name = str(ws_target.cell(row=r, column=target_name_col + 1).value or '').strip()
            if not raw_name:
                continue
            bn = norm_branch(raw_name)
            if bn and bn in edit_data:
                val_pct = edit_data[bn]
                original = ws_target.cell(row=r, column=target_aon_col + 1).value
                if original is not None and isinstance(original, (int, float)) and original <= 1.5:
                    ws_target.cell(row=r, column=target_aon_col + 1, value=round(val_pct / 100, 6))
                else:
                    ws_target.cell(row=r, column=target_aon_col + 1, value=val_pct)
                updated = True

        if updated:
            # Save as .xlsx (replace .xls)
            new_filepath = filepath.rsplit('.', 1)[0] + '.xlsx'
            wb_new.save(new_filepath)
            wb_new.close()

            # Remove old .xls, update aon_files in data.json
            if new_filepath != filepath:
                os.remove(filepath)
                data = load_data()
                old_filename = os.path.basename(filepath)
                new_filename = os.path.basename(new_filepath)
                for k, v in data.get('aon_files', {}).items():
                    if v == old_filename:
                        data['aon_files'][k] = new_filename
                save_data(data)

            return True

    wb_new.close()
    return False


# ─── API Endpoints ────────────────────────────────────────────────────────────

@app.route('/api/ping')
def api_ping():
    """ตรวจสอบว่า server ทำงานอยู่"""
    return jsonify({'ok': True, 'version': '1.0', 'timestamp': datetime.now().isoformat()})

@app.route('/api/data')
def _folder_last_modified(folder):
    """หาวันที่แก้ไขล่าสุดของไฟล์ในโฟลเดอร์"""
    latest = None
    if os.path.exists(folder):
        for f in os.listdir(folder):
            fp = os.path.join(folder, f)
            if os.path.isfile(fp):
                mtime = os.path.getmtime(fp)
                if latest is None or mtime > latest:
                    latest = mtime
    return datetime.fromtimestamp(latest).strftime('%d/%m/%Y %H:%M') if latest else None

def api_get_data():
    """คืนข้อมูลทั้งหมดที่ upload ไว้"""
    data = load_data()
    return jsonify({
        'ok': True,
        'pr': data.get('pr', {}),
        'aon': data.get('aon', {}),
        'pr_cat_names': data.get('pr_cat_names', []),
        'pr_files': data.get('pr_files', {}),
        'aon_files': data.get('aon_files', {}),
        'pr_last_modified': _folder_last_modified(PR_DIR),
        'aon_last_modified': _folder_last_modified(AON_DIR),
    })

@app.route('/api/upload/pr', methods=['POST', 'OPTIONS'])
def api_upload_pr():
    """รับ upload ไฟล์ PR (GUI_019 format)"""
    if request.method == 'OPTIONS':
        return '', 204

    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': 'ไม่ได้เลือกไฟล์'}), 400

    mode = request.form.get('mode', 'auto')
    manual_mk = request.form.get('manualMK', None)

    data = load_data()
    results = []
    errors = []

    for f in files:
        filename = f.filename
        try:
            # Save temp file
            temp_path = os.path.join(UPLOAD_DIR, f'_temp_{filename}')
            f.save(temp_path)

            # Parse
            result = parse_pr_file(temp_path, filename, manual_mk if mode == 'manual' else None)

            # Auto-rename & store
            mk = result['mk']
            ext = os.path.splitext(filename)[1] or '.xlsx'
            new_filename = f'PR_{mk}{ext}'
            dest_path = os.path.join(PR_DIR, new_filename)
            shutil.move(temp_path, dest_path)

            # Update data store
            if mk not in data['pr']:
                data['pr'][mk] = {}
            data['pr'][mk] = result['data']
            data['pr_files'][mk] = new_filename

            # Update cat_names
            for cat in result['cat_names']:
                if cat not in data.get('pr_cat_names', []):
                    data.setdefault('pr_cat_names', []).append(cat)

            results.append({
                'mk': mk,
                'count': result['count'],
                'filename': new_filename,
                'cat_names': result['cat_names'],
            })

        except Exception as e:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            errors.append({'filename': filename, 'error': str(e)})

    save_data(data)

    return jsonify({
        'ok': True,
        'results': results,
        'errors': errors,
        'pr_data': data['pr'],
        'pr_cat_names': data.get('pr_cat_names', []),
    })

@app.route('/api/upload/aon', methods=['POST', 'OPTIONS'])
def api_upload_aon():
    """รับ upload ไฟล์ AON (Always-On)"""
    if request.method == 'OPTIONS':
        return '', 204

    files = request.files.getlist('files')
    if not files:
        return jsonify({'ok': False, 'error': 'ไม่ได้เลือกไฟล์'}), 400

    mode = request.form.get('mode', 'auto')
    manual_mk = request.form.get('manualMK', None)

    data = load_data()
    results = []
    errors = []

    for f in files:
        filename = f.filename
        temp_path = os.path.join(UPLOAD_DIR, f'_temp_{filename}')
        try:
            f.save(temp_path)

            result = parse_aon_file(temp_path, filename, mode, manual_mk if mode == 'manual' else None)

            # Auto-rename & store
            months_str = '_'.join(result['processed_months']) if result['processed_months'] else 'unknown'
            ext = os.path.splitext(filename)[1] or '.xlsx'
            new_filename = f'AON_{months_str}{ext}'
            dest_path = os.path.join(AON_DIR, new_filename)
            shutil.move(temp_path, dest_path)

            # Update data store
            for mk, branch_data in result['months'].items():
                data['aon'][mk] = branch_data
                data['aon_files'][mk] = new_filename

            results.append({
                'months': result['processed_months'],
                'count': result['count'],
                'filename': new_filename,
            })

        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            errors.append({'filename': filename, 'error': str(e)})

    save_data(data)

    return jsonify({
        'ok': True,
        'results': results,
        'errors': errors,
        'aon_data': data['aon'],
    })

@app.route('/api/open-folder', methods=['POST', 'OPTIONS'])
def api_open_folder():
    """เปิดโฟลเดอร์ ข้อมูลดิบ/ ใน File Explorer"""
    if request.method == 'OPTIONS':
        return '', 204
    import subprocess, platform
    folder = UPLOAD_DIR
    try:
        if platform.system() == 'Windows':
            os.startfile(folder)
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', folder])
        else:
            subprocess.Popen(['xdg-open', folder])
        return jsonify({'ok': True, 'path': folder})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/data/pr/<mk>', methods=['DELETE', 'OPTIONS'])
def api_delete_pr(mk):
    """ลบข้อมูล PR เดือนที่ระบุ"""
    if request.method == 'OPTIONS':
        return '', 204

    data = load_data()
    deleted = False

    if mk in data['pr']:
        del data['pr'][mk]
        deleted = True
    if mk in data.get('pr_files', {}):
        # ลบไฟล์ด้วย
        filepath = os.path.join(PR_DIR, data['pr_files'][mk])
        if os.path.exists(filepath):
            os.remove(filepath)
        del data['pr_files'][mk]

    save_data(data)
    return jsonify({'ok': True, 'deleted': deleted, 'mk': mk})

@app.route('/api/data/aon/<mk>', methods=['DELETE', 'OPTIONS'])
def api_delete_aon(mk):
    """ลบข้อมูล AON เดือนที่ระบุ"""
    if request.method == 'OPTIONS':
        return '', 204

    data = load_data()
    deleted = False

    if mk in data['aon']:
        del data['aon'][mk]
        deleted = True
    if mk in data.get('aon_files', {}):
        # ลบไฟล์ด้วย
        filepath = os.path.join(AON_DIR, data['aon_files'][mk])
        if os.path.exists(filepath):
            os.remove(filepath)
        del data['aon_files'][mk]

    save_data(data)
    return jsonify({'ok': True, 'deleted': deleted, 'mk': mk})

@app.route('/api/data/edit/pr', methods=['POST', 'OPTIONS'])
def api_edit_pr():
    """แก้ไขข้อมูล PR — รับ JSON {mk, data: {branch: {...}}}"""
    if request.method == 'OPTIONS':
        return '', 204

    body = request.get_json()
    if not body or 'mk' not in body or 'data' not in body:
        return jsonify({'ok': False, 'error': 'ข้อมูลไม่ครบ (ต้องมี mk และ data)'}), 400

    mk = body['mk']
    edit_data = body['data']

    data = load_data()
    data['pr'][mk] = edit_data
    data['pr_files'].setdefault(mk, f'edited_{mk}')
    save_data(data)

    # --- Write back to Excel file ---
    excel_updated = False
    pr_filename = data['pr_files'].get(mk, '')
    pr_filepath = os.path.join(PR_DIR, pr_filename) if pr_filename else ''
    if pr_filepath and os.path.exists(pr_filepath) and openpyxl is not None:
        try:
            excel_updated = write_back_pr_excel(pr_filepath, edit_data, data.get('pr_cat_names', []))
        except Exception as e:
            print(f"[WARN] PR Excel write-back failed: {e}")

    return jsonify({'ok': True, 'mk': mk, 'excel_updated': excel_updated})

@app.route('/api/data/edit/aon', methods=['POST', 'OPTIONS'])
def api_edit_aon():
    """แก้ไขข้อมูล AON — รับ JSON {mk, data: {branch: value}}"""
    if request.method == 'OPTIONS':
        return '', 204

    body = request.get_json()
    if not body or 'mk' not in body or 'data' not in body:
        return jsonify({'ok': False, 'error': 'ข้อมูลไม่ครบ (ต้องมี mk และ data)'}), 400

    mk = body['mk']
    edit_data = body['data']

    data = load_data()
    data['aon'][mk] = edit_data
    data['aon_files'].setdefault(mk, f'edited_{mk}')
    save_data(data)

    # --- Write back to Excel file ---
    excel_updated = False
    aon_filename = data['aon_files'].get(mk, '')
    aon_filepath = os.path.join(AON_DIR, aon_filename) if aon_filename else ''
    if aon_filepath and os.path.exists(aon_filepath):
        try:
            excel_updated = write_back_aon_excel(aon_filepath, mk, edit_data)
        except Exception as e:
            print(f"[WARN] AON Excel write-back failed: {e}")

    return jsonify({'ok': True, 'mk': mk, 'excel_updated': excel_updated})

@app.route('/api/data/clear', methods=['DELETE', 'OPTIONS'])
def api_clear_all():
    """ล้างข้อมูลทั้งหมด"""
    if request.method == 'OPTIONS':
        return '', 204

    # ล้าง JSON
    save_data({'pr': {}, 'aon': {}, 'pr_files': {}, 'aon_files': {}, 'pr_cat_names': []})

    # ล้างไฟล์
    for folder in [PR_DIR, AON_DIR]:
        for f in os.listdir(folder):
            fp = os.path.join(folder, f)
            if os.path.isfile(fp):
                os.remove(fp)

    return jsonify({'ok': True, 'message': 'ล้างข้อมูลทั้งหมดเรียบร้อย'})

@app.route('/api/files')
def api_list_files():
    """รายชื่อไฟล์ที่ upload ไว้"""
    pr_files = sorted(os.listdir(PR_DIR)) if os.path.exists(PR_DIR) else []
    aon_files = sorted(os.listdir(AON_DIR)) if os.path.exists(AON_DIR) else []

    data = load_data()
    return jsonify({
        'ok': True,
        'pr_files': pr_files,
        'aon_files': aon_files,
        'pr_months': sorted(data.get('pr', {}).keys()),
        'aon_months': sorted(data.get('aon', {}).keys()),
    })

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("=" * 60)
    print("  Dashboard PR — Development Server")
    print(f"  http://localhost:{PORT}")
    print("=" * 60)
    print(f"  Base directory : {BASE_DIR}")
    print(f"  Raw data dir   : {RAW_DATA_DIR}")
    print(f"  Data file      : {DATA_FILE}")
    print()

    # ตรวจ dependencies
    missing = []
    if openpyxl is None:
        missing.append('openpyxl')
    if xlrd is None:
        missing.append('xlrd')
    if missing:
        print(f"  [WARNING] ไม่พบ library: {', '.join(missing)}")
        print(f"  ติดตั้งด้วย: pip install {' '.join(missing)}")
        print()

    app.run(host='0.0.0.0', port=PORT, debug=True)
